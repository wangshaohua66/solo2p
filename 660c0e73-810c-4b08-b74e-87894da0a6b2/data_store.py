from __future__ import annotations

import os
import sqlite3
from datetime import datetime, timedelta
from typing import Optional

from loguru import logger

from image_analyzer import AnalysisResult, AlarmLightState, GaugeType


class DataStore:
    def __init__(self, db_path: str, export_dir: str):
        self.db_path = db_path
        self.export_dir = export_dir
        os.makedirs(os.path.dirname(db_path), exist_ok=True)
        os.makedirs(export_dir, exist_ok=True)
        self._init_db()

    def _init_db(self):
        with self._get_conn() as conn:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS patrol_records (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    point_id TEXT NOT NULL,
                    point_name TEXT NOT NULL,
                    system TEXT,
                    gauge_type TEXT NOT NULL,
                    value REAL,
                    unit TEXT,
                    raw_text TEXT,
                    confidence REAL,
                    angle REAL,
                    is_anomaly INTEGER DEFAULT 0,
                    anomaly_type TEXT,
                    alarm_states TEXT,
                    timestamp TEXT NOT NULL,
                    created_at TEXT DEFAULT (datetime('now', 'localtime'))
                )
            """)
            conn.execute("""
                CREATE TABLE IF NOT EXISTS alert_events (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    point_id TEXT NOT NULL,
                    point_name TEXT NOT NULL,
                    alert_type TEXT NOT NULL,
                    alert_level TEXT NOT NULL,
                    current_value TEXT,
                    threshold_value TEXT,
                    message TEXT,
                    acknowledged INTEGER DEFAULT 0,
                    acknowledged_by TEXT,
                    acknowledged_at TEXT,
                    timestamp TEXT NOT NULL,
                    created_at TEXT DEFAULT (datetime('now', 'localtime'))
                )
            """)
            conn.execute("""
                CREATE INDEX IF NOT EXISTS idx_patrol_point_time
                ON patrol_records(point_id, timestamp)
            """)
            conn.execute("""
                CREATE INDEX IF NOT EXISTS idx_patrol_anomaly
                ON patrol_records(is_anomaly, timestamp)
            """)
            conn.execute("""
                CREATE INDEX IF NOT EXISTS idx_alert_time
                ON alert_events(timestamp)
            """)
            conn.execute("""
                CREATE INDEX IF NOT EXISTS idx_alert_ack
                ON alert_events(acknowledged)
            """)

    def _get_conn(self) -> sqlite3.Connection:
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA synchronous=NORMAL")
        return conn

    def save_patrol_record(self, result: AnalysisResult, point_config: dict):
        is_anomaly = 0
        anomaly_type = None
        value = None
        raw_text = None
        confidence = None
        angle = None
        alarm_states = None

        if result.gauge_type == GaugeType.DIGITAL and result.digital_result:
            value = result.digital_result.value
            raw_text = result.digital_result.raw_text
            confidence = result.digital_result.confidence
        elif result.gauge_type == GaugeType.POINTER and result.pointer_result:
            value = result.pointer_result.value
            angle = result.pointer_result.angle
            confidence = result.pointer_result.confidence
        elif result.gauge_type == GaugeType.ALARM_LIGHTS and result.alarm_results:
            confidence = max(
                (r.confidence for r in result.alarm_results), default=0.0,
            )
            states = []
            for ar in result.alarm_results:
                if ar.state == AlarmLightState.RED:
                    is_anomaly = 1
                    anomaly_type = anomaly_type or "alarm_red"
                    states.append(f"{ar.label}=RED")
                elif ar.state == AlarmLightState.YELLOW:
                    states.append(f"{ar.label}=YELLOW")
                else:
                    states.append(f"{ar.label}={ar.state.value}")
            alarm_states = ";".join(states)

        if result.error:
            is_anomaly = 1
            anomaly_type = "recognition_error"

        try:
            with self._get_conn() as conn:
                conn.execute("""
                    INSERT INTO patrol_records
                    (point_id, point_name, system, gauge_type, value, unit,
                     raw_text, confidence, angle, is_anomaly, anomaly_type,
                     alarm_states, timestamp)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    result.point_id,
                    point_config.get("name", result.point_id),
                    point_config.get("system", ""),
                    result.gauge_type.value,
                    value,
                    point_config.get("unit", ""),
                    raw_text,
                    confidence,
                    angle,
                    is_anomaly,
                    anomaly_type,
                    alarm_states,
                    result.timestamp,
                ))
        except Exception as e:
            logger.error(f"Failed to save patrol record: {e}")

    def save_alert_event(
        self,
        point_id: str,
        point_name: str,
        alert_type: str,
        alert_level: str,
        current_value: str,
        threshold_value: str,
        message: str,
        timestamp: str,
    ):
        try:
            with self._get_conn() as conn:
                conn.execute("""
                    INSERT INTO alert_events
                    (point_id, point_name, alert_type, alert_level,
                     current_value, threshold_value, message, timestamp)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    point_id, point_name, alert_type, alert_level,
                    current_value, threshold_value, message, timestamp,
                ))
        except Exception as e:
            logger.error(f"Failed to save alert event: {e}")

    def acknowledge_alert(self, alert_id: int, acknowledged_by: str):
        try:
            with self._get_conn() as conn:
                conn.execute("""
                    UPDATE alert_events
                    SET acknowledged = 1,
                        acknowledged_by = ?,
                        acknowledged_at = datetime('now', 'localtime')
                    WHERE id = ?
                """, (acknowledged_by, alert_id))
        except Exception as e:
            logger.error(f"Failed to acknowledge alert: {e}")

    def query_records(
        self,
        start_time: Optional[str] = None,
        end_time: Optional[str] = None,
        point_id: Optional[str] = None,
        anomaly_only: bool = False,
        limit: int = 1000,
    ) -> list[dict]:
        conditions = []
        params = []

        if start_time:
            conditions.append("timestamp >= ?")
            params.append(start_time)
        if end_time:
            conditions.append("timestamp <= ?")
            params.append(end_time)
        if point_id:
            conditions.append("point_id = ?")
            params.append(point_id)
        if anomaly_only:
            conditions.append("is_anomaly = 1")

        where = " AND ".join(conditions) if conditions else "1=1"
        query = f"""
            SELECT * FROM patrol_records
            WHERE {where}
            ORDER BY timestamp DESC
            LIMIT ?
        """
        params.append(limit)

        try:
            with self._get_conn() as conn:
                rows = conn.execute(query, params).fetchall()
                return [dict(row) for row in rows]
        except Exception as e:
            logger.error(f"Query records failed: {e}")
            return []

    def query_alerts(
        self,
        start_time: Optional[str] = None,
        end_time: Optional[str] = None,
        acknowledged: Optional[bool] = None,
        limit: int = 500,
    ) -> list[dict]:
        conditions = []
        params = []

        if start_time:
            conditions.append("timestamp >= ?")
            params.append(start_time)
        if end_time:
            conditions.append("timestamp <= ?")
            params.append(end_time)
        if acknowledged is not None:
            conditions.append("acknowledged = ?")
            params.append(1 if acknowledged else 0)

        where = " AND ".join(conditions) if conditions else "1=1"
        query = f"""
            SELECT * FROM alert_events
            WHERE {where}
            ORDER BY timestamp DESC
            LIMIT ?
        """
        params.append(limit)

        try:
            with self._get_conn() as conn:
                rows = conn.execute(query, params).fetchall()
                return [dict(row) for row in rows]
        except Exception as e:
            logger.error(f"Query alerts failed: {e}")
            return []

    def export_excel(
        self,
        start_time: Optional[str] = None,
        end_time: Optional[str] = None,
        point_id: Optional[str] = None,
        anomaly_only: bool = False,
    ) -> Optional[str]:
        try:
            from openpyxl import Workbook
            from openpyxl.chart import LineChart, Reference
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            records = self.query_records(
                start_time=start_time,
                end_time=end_time,
                point_id=point_id,
                anomaly_only=anomaly_only,
            )

            if not records:
                logger.warning("No records to export")
                return None

            wb = Workbook()

            ws_data = wb.active
            ws_data.title = "巡检数据"

            headers = [
                "时间", "点位ID", "点位名称", "系统", "仪表类型",
                "读数", "单位", "原始文本", "置信度", "角度",
                "是否异常", "异常类型", "报警灯状态",
            ]
            header_fill = PatternFill(
                start_color="4472C4", end_color="4472C4", fill_type="solid",
            )
            header_font = Font(color="FFFFFF", bold=True, size=11)
            anomaly_fill = PatternFill(
                start_color="FF0000", end_color="FF0000", fill_type="solid",
            )
            warning_fill = PatternFill(
                start_color="FFC000", end_color="FFC000", fill_type="solid",
            )
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            for col, header in enumerate(headers, 1):
                cell = ws_data.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border

            for row_idx, record in enumerate(records, 2):
                values = [
                    record.get("timestamp", ""),
                    record.get("point_id", ""),
                    record.get("point_name", ""),
                    record.get("system", ""),
                    record.get("gauge_type", ""),
                    record.get("value", ""),
                    record.get("unit", ""),
                    record.get("raw_text", ""),
                    record.get("confidence", ""),
                    record.get("angle", ""),
                    "是" if record.get("is_anomaly") else "否",
                    record.get("anomaly_type", ""),
                    record.get("alarm_states", ""),
                ]
                for col, val in enumerate(values, 1):
                    cell = ws_data.cell(row=row_idx, column=col, value=val)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center")

                if record.get("is_anomaly"):
                    for col in range(1, len(headers) + 1):
                        ws_data.cell(
                            row=row_idx, column=col,
                        ).fill = anomaly_fill
                        ws_data.cell(
                            row=row_idx, column=col,
                        ).font = Font(color="FFFFFF", bold=True)

            for col in range(1, len(headers) + 1):
                ws_data.column_dimensions[
                    get_column_letter(col)
                ].width = 16

            ws_alert = wb.create_sheet("告警记录")
            alerts = self.query_alerts(
                start_time=start_time, end_time=end_time,
            )
            alert_headers = [
                "时间", "点位ID", "点位名称", "告警类型",
                "告警级别", "当前值", "阈值", "告警信息", "已确认",
            ]
            for col, header in enumerate(alert_headers, 1):
                cell = ws_alert.cell(row=1, column=col, value=header)
                cell.fill = PatternFill(
                    start_color="C00000", end_color="C00000",
                    fill_type="solid",
                )
                cell.font = Font(color="FFFFFF", bold=True, size=11)
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border

            for row_idx, alert in enumerate(alerts, 2):
                vals = [
                    alert.get("timestamp", ""),
                    alert.get("point_id", ""),
                    alert.get("point_name", ""),
                    alert.get("alert_type", ""),
                    alert.get("alert_level", ""),
                    alert.get("current_value", ""),
                    alert.get("threshold_value", ""),
                    alert.get("message", ""),
                    "是" if alert.get("acknowledged") else "否",
                ]
                for col, val in enumerate(vals, 1):
                    cell = ws_alert.cell(row=row_idx, column=col, value=val)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="center")

            for col in range(1, len(alert_headers) + 1):
                ws_alert.column_dimensions[
                    get_column_letter(col)
                ].width = 16

            ws_chart = wb.create_sheet("趋势图")
            value_records = [
                r for r in records
                if r.get("value") is not None and r.get("gauge_type") in (
                    "digital_gauge", "pointer_gauge",
                )
            ]
            if value_records:
                chart_headers = ["时间", "点位ID", "读数"]
                for col, header in enumerate(chart_headers, 1):
                    ws_chart.cell(row=1, column=col, value=header)

                for row_idx, rec in enumerate(value_records, 2):
                    ws_chart.cell(
                        row=row_idx, column=1, value=rec.get("timestamp", ""),
                    )
                    ws_chart.cell(
                        row=row_idx, column=2, value=rec.get("point_id", ""),
                    )
                    ws_chart.cell(
                        row=row_idx, column=3, value=rec.get("value"),
                    )

                unique_points = list({
                    r["point_id"] for r in value_records
                })[:5]

                for pidx, pid in enumerate(unique_points):
                    chart = LineChart()
                    chart.title = f"{pid} 趋势"
                    chart.style = 10
                    chart.y_axis.title = "读数"
                    chart.x_axis.title = "时间"
                    chart.width = 25
                    chart.height = 15

                    start_row = 2
                    end_row = len(value_records) + 1
                    data_ref = Reference(
                        ws_chart, min_col=3,
                        min_row=start_row, max_row=end_row,
                    )
                    cats_ref = Reference(
                        ws_chart, min_col=1,
                        min_row=start_row, max_row=end_row,
                    )
                    chart.add_data(data_ref, titles_from_data=False)
                    chart.set_categories(cats_ref)
                    chart.series[0].name = pid

                    ws_chart.add_chart(
                        chart, f"E{1 + pidx * 17}",
                    )

            now_str = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"patrol_report_{now_str}.xlsx"
            filepath = os.path.join(self.export_dir, filename)
            wb.save(filepath)
            logger.info(f"Excel report exported: {filepath}")
            return filepath

        except ImportError:
            logger.error("openpyxl not installed, cannot export Excel")
            return None
        except Exception as e:
            logger.error(f"Excel export failed: {e}")
            return None

    def cleanup_old_records(self, retention_days: int = 90):
        cutoff = (
            datetime.now() - timedelta(days=retention_days)
        ).isoformat()
        try:
            with self._get_conn() as conn:
                deleted_patrol = conn.execute(
                    "DELETE FROM patrol_records WHERE timestamp < ?",
                    (cutoff,),
                ).rowcount
                deleted_alerts = conn.execute(
                    "DELETE FROM alert_events WHERE timestamp < ?",
                    (cutoff,),
                ).rowcount
                if deleted_patrol or deleted_alerts:
                    logger.info(
                        f"Cleaned up {deleted_patrol} patrol records, "
                        f"{deleted_alerts} alert events older than {cutoff}"
                    )
        except Exception as e:
            logger.error(f"Record cleanup failed: {e}")

    def get_stats(self) -> dict:
        try:
            with self._get_conn() as conn:
                total = conn.execute(
                    "SELECT COUNT(*) FROM patrol_records"
                ).fetchone()[0]
                anomalies = conn.execute(
                    "SELECT COUNT(*) FROM patrol_records WHERE is_anomaly=1"
                ).fetchone()[0]
                unacked = conn.execute(
                    "SELECT COUNT(*) FROM alert_events WHERE acknowledged=0"
                ).fetchone()[0]
                last_patrol = conn.execute(
                    "SELECT MAX(timestamp) FROM patrol_records"
                ).fetchone()[0]
                return {
                    "total_records": total,
                    "anomaly_count": anomalies,
                    "unacknowledged_alerts": unacked,
                    "last_patrol_time": last_patrol,
                }
        except Exception as e:
            logger.error(f"Get stats failed: {e}")
            return {}
