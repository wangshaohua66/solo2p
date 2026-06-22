"""
excel_parser.py
================================================================================
Excel 智能解析模块：解析工资表和人员变动表，提取结构化数据。

特性：
  1. 自动识别工资表 / 人员变动表结构，支持常见表头变体
  2. 自动定位表头所在行（跳过标题、说明、空行）
  3. 字段映射：依据 config.yaml 的 field_mapping 将本地表头映射到标准字段
  4. 自动转换日期格式、数值类型，身份证号保持文本避免精度丢失
  5. 合并单元格表头展开处理
"""

from __future__ import annotations

import logging
import os
import re
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Dict, List, Optional, Sequence, Tuple

import pandas as pd
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# 标准字段白名单
_PERSONNEL_STD: Tuple[str, ...] = (
    "id_card", "name", "insurance_type", "change_type", "change_date", "base_salary"
)
_WAGE_STD: Tuple[str, ...] = (
    "id_card", "name", "pension_base", "medical_base", "unemployment_base",
    "workinjury_base", "maternity_base"
)


@dataclass
class ParsedSheet:
    """单张工作表解析结果。"""
    sheet_name: str
    table_type: str  # "wage" / "personnel" / "unknown"
    header_row: int
    rows: List[Dict[str, Any]] = field(default_factory=list)
    column_map: Dict[str, str] = field(default_factory=dict)
    raw_columns: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)

    @property
    def row_count(self) -> int:
        return len(self.rows)


class ExcelParser:
    """Excel 智能解析器。"""

    def __init__(self, config: Dict[str, Any]) -> None:
        self.config = config
        mapping_cfg: Dict[str, Any] = config.get("field_mapping", {})
        self.personnel_aliases = self._build_alias_map(mapping_cfg.get("personnel", {}))
        self.wage_aliases = self._build_alias_map(mapping_cfg.get("wage", {}))

    # ---------------------------- 公共入口 ----------------------------

    def parse_file(self, file_path: str, sheet_name: Optional[str] = None,
                    table_hint: Optional[str] = None) -> List[ParsedSheet]:
        """解析 Excel 文件，返回所有识别到的工作表结果。

        Args:
            file_path: xlsx / xls / csv 路径
            sheet_name: 指定工作表名，None 则解析全部
            table_hint: "wage" / "personnel" 强制类型提示
        """
        if not os.path.isfile(file_path):
            raise FileNotFoundError(f"Excel 文件不存在：{file_path}")
        ext = os.path.splitext(file_path)[1].lower()
        results: List[ParsedSheet] = []
        if ext == ".csv":
            df = pd.read_csv(file_path, dtype=str, header=None, keep_default_na=False)
            results.append(self._parse_dataframe(df, os.path.basename(file_path), table_hint))
        elif ext in (".xlsx", ".xlsm", ".xltx"):
            results = self._parse_xlsx(file_path, sheet_name, table_hint)
        elif ext == ".xls":
            df = pd.read_excel(file_path, dtype=str, header=None, sheet_name=sheet_name)
            results.append(self._parse_dataframe(df, sheet_name or "Sheet1", table_hint))
        else:
            raise ValueError(f"不支持的文件类型：{ext}")
        logger.info("解析完成：%s，共 %d 张表", file_path, len(results))
        return results

    def parse_wage_table(self, file_path: str,
                          sheet_name: Optional[str] = None) -> ParsedSheet:
        """便捷方法：仅解析工资表（取第一张识别为 wage 的表）。"""
        for sheet in self.parse_file(file_path, sheet_name=sheet_name, table_hint="wage"):
            if sheet.table_type == "wage":
                return sheet
        raise ValueError(f"未在 {file_path} 中识别到工资表")

    def parse_personnel_table(self, file_path: str,
                               sheet_name: Optional[str] = None) -> ParsedSheet:
        """便捷方法：仅解析人员变动表。"""
        for sheet in self.parse_file(file_path, sheet_name=sheet_name, table_hint="personnel"):
            if sheet.table_type == "personnel":
                return sheet
        raise ValueError(f"未在 {file_path} 中识别到人员变动表")

    # ---------------------------- 内部解析 ----------------------------

    def _parse_xlsx(self, file_path: str, sheet_name: Optional[str],
                    table_hint: Optional[str]) -> List[ParsedSheet]:
        """使用 openpyxl 读取以保留合并单元格信息，再用 pandas 解析。"""
        # 非 read_only 模式：合并单元格访问需要读写权限
        wb = load_workbook(file_path, data_only=True, read_only=False)
        names = [sheet_name] if sheet_name else wb.sheetnames
        results: List[ParsedSheet] = []
        for nm in names:
            if nm not in wb.sheetnames:
                logger.warning("工作表 %s 不存在，跳过", nm)
                continue
            ws = wb[nm]
            data = self._sheet_to_list(ws)
            if not data:
                continue
            df2 = pd.DataFrame(data)
            results.append(self._parse_dataframe(df2, nm, table_hint))
        return results

    def _parse_dataframe(self, df: pd.DataFrame, sheet_name: str,
                          table_hint: Optional[str]) -> ParsedSheet:
        """定位表头行并解析 DataFrame。"""
        # 全部转为字符串，统一处理
        df = df.fillna("").astype(str)
        # 去除完全空行
        df = df[[c for c in df.columns if str(c).strip() != ""]]
        # 定位表头行
        header_row, table_type, column_map = self._detect_header(df, table_hint)
        result = ParsedSheet(sheet_name=sheet_name, table_type=table_type,
                              header_row=header_row, column_map=column_map,
                              raw_columns=list(df.iloc[header_row].astype(str)))
        if header_row is None or not column_map:
            result.warnings.append("未识别到有效表头，请检查文件格式")
            return result
        # 取表头以下数据行
        header_values = [str(v) for v in df.iloc[header_row].tolist()]
        body = df.iloc[header_row + 1:].reset_index(drop=True)
        # 用表头重命名列，便于 _row_to_record 按列名索引
        body.columns = header_values
        for idx, row in body.iterrows():
            record = self._row_to_record(row, column_map, table_type)
            if record and self._has_any_value(record):
                result.rows.append(record)
        logger.info("工作表 %s（%s）解析 %d 行数据",
                    sheet_name, table_type, len(result.rows))
        return result

    def _detect_header(self, df: pd.DataFrame, hint: Optional[str]
                       ) -> Tuple[Optional[int], str, Dict[str, str]]:
        """扫描前若干行，找出匹配字段最多的行作为表头。

        返回 (表头行号, 表类型, 标准字段->列序号映射)。
        """
        best_row: Optional[int] = None
        best_type = "unknown"
        best_map: Dict[str, str] = {}
        best_score = 0
        scan_rows = min(len(df), 20)
        for r in range(scan_rows):
            cells = [str(v).strip() for v in df.iloc[r].tolist()]
            if not any(cells):
                continue
            # 工资表探测
            wage_map = self._match_aliases(cells, self.wage_aliases)
            pers_map = self._match_aliases(cells, self.personnel_aliases)
            wage_score = len(wage_map)
            pers_score = len(pers_map)
            # 根据提示优先
            if hint == "wage" and wage_score >= 1:
                score = wage_score + 10
            elif hint == "personnel" and pers_score >= 1:
                score = pers_score + 10
            elif wage_score >= pers_score and wage_score >= 2:
                score = wage_score
            elif pers_score >= 2:
                score = pers_score
            else:
                score = max(wage_score, pers_score)
            if score > best_score:
                best_score = score
                best_row = r
                if pers_score > wage_score:
                    best_type, best_map = "personnel", pers_map
                else:
                    best_type, best_map = "wage", wage_map
        if best_score < 2:
            return None, "unknown", {}
        return best_row, best_type, best_map

    @staticmethod
    def _match_aliases(cells: Sequence[str], alias_map: Dict[str, List[str]]
                       ) -> Dict[str, str]:
        """将单元格文本与别名匹配，返回 {标准字段: 列名}。

        列名为 cells 中的原始文本（去空格），用于后续按列取值。
        """
        matched: Dict[str, str] = {}
        normalized_cells = [(i, re.sub(r"\s+", "", c)) for i, c in enumerate(cells)]
        for std, aliases in alias_map.items():
            for alias in aliases:
                norm_alias = re.sub(r"\s+", "", alias)
                if not norm_alias:
                    continue
                for col_idx, norm_cell in normalized_cells:
                    if norm_alias == norm_cell:
                        matched[std] = cells[col_idx]
                        break
                if std in matched:
                    break
                # 模糊包含（防止表头带单位/备注）
                if len(norm_alias) >= 3:
                    for col_idx, norm_cell in normalized_cells:
                        if norm_alias in norm_cell and std not in matched:
                            matched[std] = cells[col_idx]
                            break
        return matched

    @staticmethod
    def _build_alias_map(section: Dict[str, Any]) -> Dict[str, List[str]]:
        """从 config 的 field_mapping 段构建 {标准字段: [别名列表]}。"""
        result: Dict[str, List[str]] = {}
        for std_field, spec in section.items():
            if isinstance(spec, dict):
                aliases = list(spec.get("aliases", []))
                std = spec.get("standard", std_field)
            else:
                aliases = []
                std = std_field
            aliases = [a for a in aliases if a]
            # 标准字段名也作为别名
            if std not in aliases:
                aliases.append(std)
            result[std] = aliases
        return result

    def _row_to_record(self, row: pd.Series, column_map: Dict[str, str],
                       table_type: str) -> Dict[str, Any]:
        """将一行 Series 按 column_map 转为标准字段记录，并做类型转换。"""
        record: Dict[str, Any] = {}
        # 建立列名->值查找
        row_dict = {str(k): v for k, v in row.items()}
        for std_field, col_name in column_map.items():
            value = row_dict.get(col_name, "")
            value = self._clean_cell(value)
            record[std_field] = self._convert_value(value, std_field)
        return record

    @staticmethod
    def _clean_cell(value: Any) -> str:
        if value is None:
            return ""
        text = str(value).strip()
        # 去除常见全角空格与不可见字符
        text = text.replace("\u3000", "").replace("\xa0", "")
        # 去除首尾引号
        if len(text) >= 2 and text[0] in "'\"" and text[-1] == text[0]:
            text = text[1:-1]
        return text.strip()

    def _convert_value(self, text: str, field: str) -> Any:
        """按字段类型转换：身份证号→文本，基数→浮点，日期→标准化字符串。"""
        if not text:
            return ""
        if field == "id_card":
            # 强制文本，去除可能的小数
            if text.endswith(".0"):
                text = text[:-2]
            return text.upper().replace(" ", "")
        if field == "name":
            return text
        if field.endswith("_base") or field == "base_salary":
            return self._to_number(text)
        if field in ("change_date",):
            return self._normalize_date(text)
        if field in ("change_type", "insurance_type"):
            return text
        return text

    @staticmethod
    def _to_number(text: str) -> Optional[float]:
        cleaned = text.replace(",", "").replace("¥", "").replace("￥", "").replace("元", "").strip()
        if not cleaned:
            return None
        try:
            return float(cleaned)
        except ValueError:
            logger.debug("无法转换为数值：%r", text)
            return None

    @staticmethod
    def _normalize_date(text: str) -> str:
        """将多种日期格式统一为 YYYY-MM-DD。"""
        if isinstance(text, (datetime,)):
            return text.strftime("%Y-%m-%d")
        text = text.strip()
        # Excel 序列号
        if re.match(r"^\d{4,5}$", text):
            try:
                serial = int(text)
                if 1 <= serial <= 60000:
                    return (datetime(1899, 12, 30) + pd.Timedelta(days=serial)).strftime("%Y-%m-%d")
            except (ValueError, OverflowError):
                pass
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d", "%Y年%m月%d日", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(text, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
        return text

    @staticmethod
    def _has_any_value(record: Dict[str, Any]) -> bool:
        return any(v not in (None, "", []) for v in record.values())

    @staticmethod
    def _sheet_to_list(ws: Any) -> List[List[Any]]:
        """将 openpyxl worksheet 转为二维列表，填充合并单元格首值。"""
        data: List[List[Any]] = []
        merged_ranges = list(ws.merged_cells.ranges)
        for row in ws.iter_rows(values_only=True):
            data.append(list(row))
        # 填充合并单元格
        for rng in merged_ranges:
            try:
                min_col, min_row, max_col, max_row = rng.bounds
                val = ws.cell(row=min_row, column=min_col).value
                for r in range(min_row - 1, max_row):
                    if r >= len(data):
                        continue
                    for c in range(min_col - 1, max_col):
                        if c < len(data[r]):
                            data[r][c] = val
            except Exception:  # noqa: BLE001
                continue
        # 补齐列宽
        if data:
            width = max(len(r) for r in data)
            for r in data:
                while len(r) < width:
                    r.append(None)
        return data


def to_records(sheet: ParsedSheet) -> List[Dict[str, Any]]:
    """便捷函数：将 ParsedSheet 转为纯字典列表。"""
    return list(sheet.rows)
