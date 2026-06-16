from typing import List, Dict
from io import BytesIO

import openpyxl

from ..utils import (
    logger,
    download_file,
    safe_float,
    today_str,
    col_letter_to_index,
    Color,
    color_text,
)


class ExcelSpider:
    """采集可下载Excel附件的爬虫"""

    spider_type = "excel"

    def __init__(self):
        self.name = "ExcelSpider"

    def crawl(self, market_info: Dict) -> List[Dict]:
        market_id = market_info["id"]
        market_name = market_info.get("name", market_id)
        url_template = market_info["url"]
        date_format = market_info.get("date_format", "%Y%m%d")
        sheet_name = market_info.get("sheet_name")
        start_row = int(market_info.get("start_row", 1))
        column_mapping = market_info.get("column_mapping", {})

        today = today_str()
        date_str = today_str(fmt=date_format)
        url = url_template.format(date=date_str)

        logger.info(color_text(f"[Excel] 开始采集 {market_name} ({market_id}) URL={url}", Color.CYAN))

        file_bytes = None
        try:
            file_bytes = download_file(url)
        except Exception as e:
            logger.warning(color_text(f"[Excel] {market_name} 今日文件下载失败，尝试昨日: {e}", Color.YELLOW))
            for offset in (1, 2, 3):
                try:
                    from ..utils import date_days_ago
                    alt_date = date_days_ago(offset, fmt=date_format)
                    alt_url = url_template.format(date=alt_date)
                    file_bytes = download_file(alt_url)
                    logger.info(f"[Excel] 使用{offset}天前的文件: {alt_url}")
                    break
                except Exception as e2:
                    logger.debug(f"[Excel] {offset}天前文件也失败: {e2}")
                    continue

        if file_bytes is None:
            logger.error(color_text(f"[Excel] {market_name} 连续多日文件下载失败", Color.RED))
            raise RuntimeError(f"Excel文件下载失败: {url}")

        try:
            raw_records = self._parse_excel(
                file_bytes, sheet_name, start_row, column_mapping
            )
            logger.info(
                color_text(
                    f"[Excel] {market_name} 解析完成，共{len(raw_records)}条原始记录",
                    Color.GREEN,
                )
            )
            return raw_records
        except Exception as e:
            logger.error(color_text(f"[Excel] {market_name} 解析失败: {e}", Color.RED))
            raise

    def _parse_excel(
        self,
        file_bytes: BytesIO,
        sheet_name=None,
        start_row: int = 1,
        column_mapping: Dict = None,
    ) -> List[Dict]:
        if column_mapping is None:
            column_mapping = {}
        records = []

        try:
            wb = openpyxl.load_workbook(file_bytes, data_only=True, read_only=True)
        except Exception as e:
            logger.error(f"Excel打开失败: {e}")
            return []

        try:
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.active

            if start_row < 1:
                start_row = 1

            first_row = None
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if row_idx == 1 and first_row is None:
                    first_row = [str(c).strip() if c is not None else "" for c in row]

                if row_idx < start_row:
                    continue
                if row is None or all(c is None or c == "" for c in row):
                    continue

                try:
                    record = self._build_row_record(row, column_mapping, first_row)
                    if record and str(record.get("category_name", "")).strip():
                        for price_key in ("max_price", "min_price", "avg_price"):
                            if price_key in record:
                                record[price_key] = safe_float(record[price_key])
                        if record.get("avg_price", 0) > 0 or record.get("max_price", 0) > 0:
                            records.append(record)
                except Exception as e:
                    logger.debug(f"解析第{row_idx}行失败: {e}")
                    continue
        finally:
            try:
                wb.close()
            except Exception:
                pass

        return records

    def _build_row_record(self, row, column_mapping: Dict, first_row=None) -> Dict:
        record = {}

        has_letter_keys = any(k.upper().isalpha() for k in column_mapping.keys())
        if has_letter_keys:
            for col_key, target_key in column_mapping.items():
                if col_key.upper().isalpha():
                    idx = col_letter_to_index(col_key)
                    if idx < len(row):
                        val = row[idx]
                        if val is not None:
                            record[target_key] = val
                elif first_row and col_key in first_row:
                    idx = first_row.index(col_key)
                    if idx < len(row):
                        val = row[idx]
                        if val is not None:
                            record[target_key] = val
        else:
            for col_key, target_key in column_mapping.items():
                if first_row and col_key in first_row:
                    idx = first_row.index(col_key)
                    if idx < len(row):
                        val = row[idx]
                        if val is not None:
                            record[target_key] = val

        return record
