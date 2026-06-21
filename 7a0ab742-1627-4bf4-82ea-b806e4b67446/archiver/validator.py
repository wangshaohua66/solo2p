import os
import re
import mimetypes
from datetime import datetime
from pathlib import Path
from dataclasses import dataclass, field
from typing import List, Dict, Any, Optional, Tuple
from enum import Enum

try:
    import openpyxl
except ImportError:
    openpyxl = None


class ValidationSeverity(str, Enum):
    ERROR = "error"
    WARNING = "warning"
    INFO = "info"


class ValidationType(str, Enum):
    COMPLETENESS = "completeness"
    FORMAT = "format"
    LOGIC = "logic"
    FILE = "file"


@dataclass
class ValidationIssue:
    field: str
    message: str
    severity: ValidationSeverity
    type: ValidationType
    value: Optional[Any] = None

    def to_dict(self) -> Dict[str, Any]:
        return {
            "field": self.field,
            "message": self.message,
            "severity": self.severity.value,
            "type": self.type.value,
            "value": str(self.value) if self.value is not None else None,
        }


@dataclass
class ValidationResult:
    archive_id: str
    passed: bool = True
    issues: List[ValidationIssue] = field(default_factory=list)
    metadata: Dict[str, Any] = field(default_factory=dict)

    def add_issue(self, issue: ValidationIssue):
        self.issues.append(issue)
        if issue.severity == ValidationSeverity.ERROR:
            self.passed = False

    def to_dict(self) -> Dict[str, Any]:
        return {
            "archive_id": self.archive_id,
            "passed": self.passed,
            "issues": [issue.to_dict() for issue in self.issues],
            "metadata": self.metadata,
            "error_count": len([i for i in self.issues if i.severity == ValidationSeverity.ERROR]),
            "warning_count": len([i for i in self.issues if i.severity == ValidationSeverity.WARNING]),
        }


@dataclass
class ValidationConfig:
    required_fields: List[str] = field(default_factory=lambda: [
        "title", "author", "created_date", "archive_number", "retention_period"
    ])
    date_format: str = "%Y-%m-%d"
    retention_periods: List[str] = field(default_factory=lambda: [
        "永久", "30年", "10年", "5年", "3年"
    ])
    secrecy_levels: List[str] = field(default_factory=lambda: [
        "公开", "内部", "秘密", "机密", "绝密"
    ])
    archive_number_pattern: str = r"^[A-Z0-9]+-[A-Z0-9]+-[A-Z0-9]+-[A-Z0-9]+$"
    allowed_extensions: List[str] = field(default_factory=lambda: [
        ".pdf", ".ofd", ".docx", ".xlsx", ".jpg", ".jpeg", ".png", ".tiff", ".tif"
    ])
    max_file_size_mb: int = 500
    max_batch_size: int = 100000
    enable_gbt18894: bool = True


class ArchiveValidator:
    def __init__(self, config: Optional[ValidationConfig] = None, logger=None):
        self.config = config or ValidationConfig()
        self.logger = logger

    def validate_batch(self, archives: List[Dict[str, Any]], file_dir: Optional[str] = None) -> List[ValidationResult]:
        results = []
        for archive in archives:
            file_path = None
            if file_dir and "file_name" in archive:
                file_path = os.path.join(file_dir, archive["file_name"])
            elif file_dir and "file_path" in archive:
                file_path = os.path.join(file_dir, archive["file_path"])
            
            result = self.validate_archive(archive, file_path)
            results.append(result)

        return results

    def _validate_completeness(self, metadata: Dict[str, Any], result: ValidationResult):
        for field in self.config.required_fields:
            value = metadata.get(field)
            if value is None or (isinstance(value, str) and value.strip() == ""):
                result.add_issue(ValidationIssue(
                    field=field,
                    message=f"必填字段 '{field}' 缺失或为空",
                    severity=ValidationSeverity.ERROR,
                    type=ValidationType.COMPLETENESS,
                ))

    def _validate_format(self, metadata: Dict[str, Any], result: ValidationResult):
        if "created_date" in metadata and metadata["created_date"]:
            self._validate_date_format("created_date", metadata["created_date"], result)

        if "archived_date" in metadata and metadata["archived_date"]:
            self._validate_date_format("archived_date", metadata["archived_date"], result)

        if "archive_number" in metadata and metadata["archive_number"]:
            self._validate_archive_number(metadata["archive_number"], result)

        if "retention_period" in metadata and metadata["retention_period"]:
            self._validate_retention_period(metadata["retention_period"], result)

        if "secrecy_level" in metadata and metadata["secrecy_level"]:
            self._validate_secrecy_level(metadata["secrecy_level"], result)

        if "file_size" in metadata and metadata["file_size"]:
            self._validate_file_size_format(metadata["file_size"], result)

    def _validate_logic(self, metadata: Dict[str, Any], result: ValidationResult):
        created_date = metadata.get("created_date")
        archived_date = metadata.get("archived_date")

        if created_date and archived_date:
            try:
                created_dt = datetime.strptime(str(created_date), self.config.date_format)
                archived_dt = datetime.strptime(str(archived_date), self.config.date_format)
                if created_dt > archived_dt:
                    result.add_issue(ValidationIssue(
                        field="created_date",
                        message="形成时间不能晚于归档时间",
                        severity=ValidationSeverity.ERROR,
                        type=ValidationType.LOGIC,
                        value=f"{created_date} > {archived_date}",
                    ))
            except (ValueError, TypeError):
                pass

        retention_period = metadata.get("retention_period", "")
        secrecy_level = metadata.get("secrecy_level", "")

        if retention_period and secrecy_level:
            if secrecy_level in ["机密", "绝密"] and retention_period in ["3年", "5年"]:
                result.add_issue(ValidationIssue(
                    field="retention_period",
                    message=f"保管期限 '{retention_period}' 与密级 '{secrecy_level}' 不匹配",
                    severity=ValidationSeverity.WARNING,
                    type=ValidationType.LOGIC,
                ))

    def _validate_file(self, file_path: str, metadata: Dict[str, Any], result: ValidationResult):
        path = Path(file_path)

        if not path.exists():
            result.add_issue(ValidationIssue(
                field="file_path",
                message=f"文件不存在: {file_path}",
                severity=ValidationSeverity.ERROR,
                type=ValidationType.FILE,
                value=file_path,
            ))
            return

        if not path.is_file():
            result.add_issue(ValidationIssue(
                field="file_path",
                message=f"路径不是文件: {file_path}",
                severity=ValidationSeverity.ERROR,
                type=ValidationType.FILE,
                value=file_path,
            ))
            return

        ext = path.suffix.lower()
        if ext not in self.config.allowed_extensions:
            result.add_issue(ValidationIssue(
                field="file_extension",
                message=f"不支持的文件格式: {ext}",
                severity=ValidationSeverity.WARNING,
                type=ValidationType.FILE,
                value=ext,
            ))

        actual_size = path.stat().st_size
        if "file_size" in metadata and metadata["file_size"]:
            try:
                expected_size = int(metadata["file_size"])
                if abs(actual_size - expected_size) > 100:
                    result.add_issue(ValidationIssue(
                        field="file_size",
                        message=f"文件大小与元数据记录不一致，实际: {actual_size} 字节, 记录: {expected_size} 字节",
                        severity=ValidationSeverity.ERROR,
                        type=ValidationType.FILE,
                        value=f"实际:{actual_size}, 记录:{expected_size}",
                    ))
            except (ValueError, TypeError):
                pass

        max_size_bytes = self.config.max_file_size_mb * 1024 * 1024
        if actual_size > max_size_bytes:
            result.add_issue(ValidationIssue(
                field="file_size",
                message=f"文件大小超过限制: {actual_size / (1024*1024):.2f}MB > {self.config.max_file_size_mb}MB",
                severity=ValidationSeverity.WARNING,
                type=ValidationType.FILE,
                value=actual_size,
            ))

        if ext in [".jpg", ".jpeg", ".png", ".tiff", ".tif", ".pdf",
                   ".docx", ".xlsx", ".ofd"]:
            self._check_file_readable(path, ext, result)

    def _validate_date_format(self, field: str, value: str, result: ValidationResult):
        try:
            datetime.strptime(str(value), self.config.date_format)
        except (ValueError, TypeError):
            result.add_issue(ValidationIssue(
                field=field,
                message=f"日期格式不正确，应为 {self.config.date_format}",
                severity=ValidationSeverity.ERROR,
                type=ValidationType.FORMAT,
                value=value,
            ))

    def _validate_archive_number(self, archive_number: str, result: ValidationResult):
        if not re.match(self.config.archive_number_pattern, str(archive_number)):
            result.add_issue(ValidationIssue(
                field="archive_number",
                message=f"档号格式不正确，应符合模式: {self.config.archive_number_pattern}",
                severity=ValidationSeverity.ERROR,
                type=ValidationType.FORMAT,
                value=archive_number,
            ))

    def _validate_retention_period(self, period: str, result: ValidationResult):
        if period not in self.config.retention_periods:
            result.add_issue(ValidationIssue(
                field="retention_period",
                message=f"保管期限值无效，有效值为: {', '.join(self.config.retention_periods)}",
                severity=ValidationSeverity.ERROR,
                type=ValidationType.FORMAT,
                value=period,
            ))

    def _validate_secrecy_level(self, level: str, result: ValidationResult):
        if level not in self.config.secrecy_levels:
            result.add_issue(ValidationIssue(
                field="secrecy_level",
                message=f"密级值无效，有效值为: {', '.join(self.config.secrecy_levels)}",
                severity=ValidationSeverity.ERROR,
                type=ValidationType.FORMAT,
                value=level,
            ))

    def _validate_file_size_format(self, size: Any, result: ValidationResult):
        try:
            int(size)
        except (ValueError, TypeError):
            result.add_issue(ValidationIssue(
                field="file_size",
                message="文件大小应为数字（字节数）",
                severity=ValidationSeverity.ERROR,
                type=ValidationType.FORMAT,
                value=size,
            ))

    def _check_file_readable(self, path: Path, ext: str, result: ValidationResult):
        try:
            if ext in [".jpg", ".jpeg", ".png", ".tiff", ".tif"]:
                from PIL import Image
                with Image.open(path) as img:
                    img.verify()
            elif ext == ".pdf":
                import pdfplumber
                with pdfplumber.open(str(path)) as pdf:
                    _ = len(pdf.pages)
            elif ext == ".docx":
                self._check_docx_readable(path)
            elif ext == ".xlsx":
                self._check_xlsx_readable(path)
            elif ext == ".ofd":
                self._check_ofd_readable(path)
        except Exception as e:
            result.add_issue(ValidationIssue(
                field="file_readable",
                message=f"文件无法正常读取: {str(e)}",
                severity=ValidationSeverity.ERROR,
                type=ValidationType.FILE,
                value=str(e),
            ))

    def _check_docx_readable(self, path: Path):
        import zipfile
        from xml.etree import ElementTree as ET

        with zipfile.ZipFile(str(path), "r") as zf:
            required_files = ["word/document.xml", "word/_rels/document.xml.rels", "[Content_Types].xml"]
            for f in required_files:
                if f not in zf.namelist():
                    raise ValueError(f"缺少必要文件缺失: {f}")

            with zf.open("word/document.xml") as f:
                content = f.read()
                if not content:
                    raise ValueError("document.xml 内容为空")

                try:
                    ET.fromstring(content)
                except ET.ParseError as e:
                    raise ValueError(f"XML 解析失败: {e}")

    def _check_xlsx_readable(self, path: Path):
        if openpyxl is None:
            try:
                from openpyxl import load_workbook
            except ImportError:
                raise ValueError("需要 openpyxl 库来检测 XLSX 文件")
        else:
            from openpyxl import load_workbook

        wb = load_workbook(str(path), read_only=True, data_only=True)
        if not wb.sheetnames:
            raise ValueError("工作簿中没有工作表")

        ws = wb[wb.sheetnames[0]]
        _ = ws.max_row
        wb.close()

    def _check_ofd_readable(self, path: Path):
        import zipfile
        from xml.etree import ElementTree as ET

        with zipfile.ZipFile(str(path), "r") as zf:
            required_files = ["OFD.xml", "Doc_0/DocumentRes.xml"]

            found_any = False
            for f in zf.namelist():
                if f.startswith("Doc_") and f.endswith("Page.xml"):
                    found_any = True
                    break

            if not found_any:
                raise ValueError("未找到文档页面文件")

            with zf.open("OFD.xml") as f:
                content = f.read()
                if not content:
                    raise ValueError("OFD.xml 内容为空")

                try:
                    ET.fromstring(content)
                except ET.ParseError as e:
                    raise ValueError(f"XML 解析失败: {e}")

    def check_duplicate_archive_numbers(self, archives: List[Dict[str, Any]]) -> List[ValidationIssue]:
        issues = []
        number_counts: Dict[str, int] = {}

        for archive in archives:
            archive_number = archive.get("archive_number", "")
            if archive_number:
                number_counts[archive_number] = number_counts.get(archive_number, 0) + 1

        for number, count in number_counts.items():
            if count > 1:
                issues.append(ValidationIssue(
                    field="archive_number",
                    message=f"档号重复: {number} 出现 {count} 次",
                    severity=ValidationSeverity.ERROR,
                    type=ValidationType.LOGIC,
                    value=f"{number}: {count}次",
                ))

        return issues

    def check_batch_size(self, archives: List[Dict[str, Any]]) -> Tuple[bool, str]:
        count = len(archives)
        if count > self.config.max_batch_size:
            message = (f"批次档案数量 {count} 超过最大限制 {self.config.max_batch_size} 件，"
                      f"请分批处理")
            return False, message
        return True, ""

    def validate_gbt18894(self, metadata: Dict[str, Any], result: ValidationResult):
        if not self.config.enable_gbt18894:
            return

        if "file_format" not in metadata or not metadata["file_format"]:
            result.add_issue(ValidationIssue(
                field="file_format",
                message="[GB/T 18894] 电子文件格式信息缺失",
                severity=ValidationSeverity.ERROR,
                type=ValidationType.FORMAT,
            ))

        if "file_size" not in metadata or not metadata["file_size"]:
            result.add_issue(ValidationIssue(
                field="file_size",
                message="[GB/T 18894] 电子文件大小信息缺失",
                severity=ValidationSeverity.ERROR,
                type=ValidationType.COMPLETENESS,
            ))

        if "title" in metadata and metadata["title"]:
            title = str(metadata["title"])
            if len(title) > 256:
                result.add_issue(ValidationIssue(
                    field="title",
                    message=f"[GB/T 18894] 题名长度 {len(title)} 字符，建议不超过256字符",
                    severity=ValidationSeverity.WARNING,
                    type=ValidationType.FORMAT,
                    value=f"{len(title)}字符",
                ))

        if "page_count" in metadata and metadata["page_count"]:
            try:
                pages = int(metadata["page_count"])
                if pages <= 0:
                    result.add_issue(ValidationIssue(
                        field="page_count",
                        message="[GB/T 18894] 页数应为正整数",
                        severity=ValidationSeverity.ERROR,
                        type=ValidationType.FORMAT,
                        value=metadata["page_count"],
                    ))
            except (ValueError, TypeError):
                result.add_issue(ValidationIssue(
                    field="page_count",
                    message="[GB/T 18894] 页数字段格式不正确",
                    severity=ValidationSeverity.ERROR,
                    type=ValidationType.FORMAT,
                    value=metadata["page_count"],
                ))

        if "file_name" in metadata and metadata["file_name"]:
            file_name = str(metadata["file_name"])
            forbidden_chars = '<>:"/\\|?*'
            invalid_chars = [c for c in forbidden_chars if c in file_name]
            if invalid_chars:
                result.add_issue(ValidationIssue(
                    field="file_name",
                    message=f"[GB/T 18894] 文件名包含非法字符: {''.join(invalid_chars)}",
                    severity=ValidationSeverity.WARNING,
                    type=ValidationType.FORMAT,
                    value=file_name,
                ))

        if "keywords" in metadata and metadata["keywords"]:
            keywords = str(metadata["keywords"])
            if len(keywords) > 500:
                result.add_issue(ValidationIssue(
                    field="keywords",
                    message=f"[GB/T 18894] 主题词长度 {len(keywords)} 字符，建议不超过500字符",
                    severity=ValidationSeverity.WARNING,
                    type=ValidationType.FORMAT,
                    value=f"{len(keywords)}字符",
                ))

        if "summary" in metadata and metadata["summary"]:
            summary = str(metadata["summary"])
            if len(summary) > 2000:
                result.add_issue(ValidationIssue(
                    field="summary",
                    message=f"[GB/T 18894] 摘要长度 {len(summary)} 字符，建议不超过2000字符",
                    severity=ValidationSeverity.WARNING,
                    type=ValidationType.FORMAT,
                    value=f"{len(summary)}字符",
                ))

    def validate_archive(self, metadata: Dict[str, Any], file_path: Optional[str] = None) -> ValidationResult:
        archive_id = metadata.get("archive_number", metadata.get("id", "unknown"))
        result = ValidationResult(archive_id=archive_id, metadata=metadata)

        self._validate_completeness(metadata, result)
        self._validate_format(metadata, result)
        self._validate_logic(metadata, result)
        self.validate_gbt18894(metadata, result)

        if file_path:
            self._validate_file(file_path, metadata, result)

        if self.logger:
            status = "PASSED" if result.passed else "FAILED"
            self.logger.info(
                f"Archive validation {status}: {archive_id}",
                operation_type="validation",
                obj=archive_id,
            )

        return result
