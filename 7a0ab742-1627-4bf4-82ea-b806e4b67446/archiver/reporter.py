import os
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional
from collections import Counter

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
except ImportError:
    SimpleDocTemplate = None


class ReportGenerator:
    def __init__(self, logger=None):
        self.logger = logger
        self._register_fonts()

    def _register_fonts(self):
        try:
            font_paths = [
                "/System/Library/Fonts/PingFang.ttc",
                "/System/Library/Fonts/STHeiti Light.ttc",
                "/System/Library/Fonts/Helvetica.ttc",
                "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            ]
            for font_path in font_paths:
                if os.path.exists(font_path):
                    try:
                        if SimpleDocTemplate:
                            pdfmetrics.registerFont(TTFont("ChineseFont", font_path))
                        break
                    except Exception:
                        continue
        except Exception:
            pass

    def generate_transfer_list(self, archives: List[Dict[str, Any]],
                               validation_results: Optional[List] = None,
                               output_path: str = "移交清单.xlsx",
                               org_name: str = "") -> str:
        if Workbook is None:
            raise ImportError("需要安装 openpyxl 库来生成 Excel 清单")

        wb = Workbook()

        self._create_summary_sheet(wb, archives, validation_results, org_name)
        self._create_detail_sheet(wb, archives, validation_results)
        self._create_issue_sheet(wb, validation_results)

        wb.save(output_path)

        if self.logger:
            self.logger.info(
                f"生成移交清单: {output_path}, 共 {len(archives)} 件档案",
                operation_type="generate_transfer_list",
                obj=output_path,
            )

        return output_path

    def _create_summary_sheet(self, wb, archives, validation_results, org_name):
        ws = wb.active
        ws.title = "汇总信息"

        title_font = Font(name="微软雅黑", size=16, bold=True)
        header_font = Font(name="微软雅黑", size=11, bold=True)
        normal_font = Font(name="微软雅黑", size=10)
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        fill_header = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        white_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")

        ws.merge_cells("A1:F1")
        ws["A1"] = "电子档案移交清单"
        ws["A1"].font = title_font
        ws["A1"].alignment = center_align

        summary_data = []

        summary_data.append(["移交单位", org_name or ""])
        summary_data.append(["移交时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        summary_data.append(["档案总数", len(archives)])

        type_counter = Counter()
        for archive in archives:
            cat = archive.get("category", archive.get("file_format", "未知"))
            type_counter[cat] += 1
        summary_data.append(["档案类型", ", ".join([f"{k}:{v}件" for k, v in type_counter.items()])])

        date_range = self._get_date_range(archives)
        summary_data.append(["时间范围", date_range])

        if validation_results:
            passed = sum(1 for r in validation_results if r.passed)
            failed = len(validation_results) - passed
            summary_data.append(["校验通过", f"{passed} 件"])
            summary_data.append(["校验未通过", f"{failed} 件"])
            summary_data.append(["通过率", f"{passed/len(validation_results)*100:.1f}%" if validation_results else "0%"])

        start_row = 3
        for i, (label, value) in enumerate(summary_data):
            row = start_row + i
            ws.cell(row=row, column=1, value=label).font = header_font
            ws.cell(row=row, column=1).alignment = left_align
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
            ws.cell(row=row, column=2, value=value).font = normal_font
            ws.cell(row=row, column=2).alignment = left_align

        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 60

    def _create_detail_sheet(self, wb, archives, validation_results):
        ws = wb.create_sheet("档案明细")

        header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
        normal_font = Font(name="微软雅黑", size=10)
        center_align = Alignment(horizontal="center", vertical="center")
        fill_header = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        headers = ["序号", "档号", "题名", "责任者", "形成时间", "保管期限", "密级", "文件格式", "校验状态"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = center_align
            cell.fill = fill_header

        result_map = {}
        if validation_results:
            for result in validation_results:
                result_map[result.archive_id] = result

        for i, archive in enumerate(archives, 1):
            row = i + 1
            archive_number = archive.get("archive_number", "")
            result = result_map.get(archive_number)

            ws.cell(row=row, column=1, value=i).font = normal_font
            ws.cell(row=row, column=2, value=archive_number).font = normal_font
            ws.cell(row=row, column=3, value=archive.get("title", "")).font = normal_font
            ws.cell(row=row, column=4, value=archive.get("author", "")).font = normal_font
            ws.cell(row=row, column=5, value=archive.get("created_date", "")).font = normal_font
            ws.cell(row=row, column=6, value=archive.get("retention_period", "")).font = normal_font
            ws.cell(row=row, column=7, value=archive.get("secrecy_level", "")).font = normal_font
            ws.cell(row=row, column=8, value=archive.get("file_format", "")).font = normal_font

            status = "通过" if result is None else ("通过" if result.passed else "不通过")
            status_cell = ws.cell(row=row, column=9, value=status)
            status_cell.font = normal_font
            if status == "通过":
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            else:
                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        widths = [8, 25, 40, 15, 15, 12, 10, 12, 10]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    def _create_issue_sheet(self, wb, validation_results):
        if not validation_results:
            return

        ws = wb.create_sheet("问题明细")

        header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
        normal_font = Font(name="微软雅黑", size=10)
        center_align = Alignment(horizontal="center", vertical="center")
        fill_header = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        headers = ["档号", "问题字段", "问题描述", "严重程度", "问题类型"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = center_align
            cell.fill = fill_header

        row = 2
        for result in validation_results:
            for issue in result.issues:
                ws.cell(row=row, column=1, value=result.archive_id).font = normal_font
                ws.cell(row=row, column=2, value=issue.field).font = normal_font
                ws.cell(row=row, column=3, value=issue.message).font = normal_font

                severity_cell = ws.cell(row=row, column=4, value=issue.severity.value)
                severity_cell.font = normal_font
                if issue.severity.value == "error":
                    severity_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                elif issue.severity.value == "warning":
                    severity_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

                ws.cell(row=row, column=5, value=issue.type.value).font = normal_font
                row += 1

        widths = [25, 15, 50, 12, 15]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

    def _get_date_range(self, archives):
        dates = []
        for archive in archives:
            date_str = archive.get("created_date", "")
            if date_str:
                try:
                    dates.append(datetime.strptime(str(date_str), "%Y-%m-%d"))
                except ValueError:
                    pass

        if dates:
            min_date = min(dates).strftime("%Y-%m-%d")
            max_date = max(dates).strftime("%Y-%m-%d")
            return f"{min_date} 至 {max_date}"
        return "未知"

    def generate_receipt_pdf(self, archives: List[Dict[str, Any]],
                             validation_results: Optional[List] = None,
                             output_path: str = "接收回执.pdf",
                             org_name: str = "",
                             receiver: str = "") -> str:
        if SimpleDocTemplate is None:
            raise ImportError("需要安装 reportlab 库来生成 PDF 回执")

        doc = SimpleDocTemplate(
            output_path,
            pagesize=A4,
            rightMargin=2 * cm,
            leftMargin=2 * cm,
            topMargin=2 * cm,
            bottomMargin=2 * cm,
        )

        styles = getSampleStyleSheet()
        story = []

        title_style = ParagraphStyle(
            "Title",
            parent=styles["Title"],
            fontName="ChineseFont" if self._has_chinese_font() else "Helvetica",
            fontSize=18,
            alignment=TA_CENTER,
            spaceAfter=20,
        )

        heading_style = ParagraphStyle(
            "Heading",
            parent=styles["Heading2"],
            fontName="ChineseFont" if self._has_chinese_font() else "Helvetica",
            fontSize=14,
            spaceBefore=10,
            spaceAfter=10,
        )

        normal_style = ParagraphStyle(
            "Normal",
            parent=styles["Normal"],
            fontName="ChineseFont" if self._has_chinese_font() else "Helvetica",
            fontSize=11,
            leading=18,
        )

        story.append(Paragraph("电子档案接收回执", title_style))
        story.append(Spacer(1, 0.5 * cm))

        passed = 0
        failed = 0
        if validation_results:
            passed = sum(1 for r in validation_results if r.passed)
            failed = len(validation_results) - passed
        else:
            passed = len(archives)

        receipt_info = [
            ["移交单位:", org_name or "__________"],
            ["移交时间:", datetime.now().strftime("%Y年%m月%d日")],
            ["接收单位:", "__________档案局"],
            ["档案总数:", f"{len(archives)} 件"],
            ["合格数量:", f"{passed} 件"],
            ["不合格数量:", f"{failed} 件"],
            ["接收状态:", "已接收" if failed == 0 else "部分接收"],
        ]

        info_table_data = []
        for row in receipt_info:
            info_table_data.append([Paragraph(row[0], normal_style), Paragraph(row[1], normal_style)])

        info_table = Table(info_table_data, colWidths=[3 * cm, 11 * cm])
        info_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), "ChineseFont" if self._has_chinese_font() else "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 11),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
            ("LINEBELOW", (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ]))
        story.append(info_table)
        story.append(Spacer(1, 0.8 * cm))

        story.append(Paragraph("一、档案类型统计", heading_style))
        type_counter = Counter()
        for archive in archives:
            cat = archive.get("category", archive.get("file_format", "其他"))
            type_counter[cat] += 1

        type_data = [["序号", "档案类型", "数量(件)", "占比"]]
        for i, (cat, count) in enumerate(type_counter.items(), 1):
            type_data.append([
                str(i),
                cat,
                str(count),
                f"{count/len(archives)*100:.1f}%" if archives else "0%",
            ])

        type_table = Table(type_data, colWidths=[2 * cm, 6 * cm, 3 * cm, 3 * cm])
        type_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), "ChineseFont" if self._has_chinese_font() else "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightblue),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
        ]))
        story.append(type_table)
        story.append(Spacer(1, 0.8 * cm))

        if validation_results and failed > 0:
            story.append(Paragraph("二、不合格件明细", heading_style))

            issue_data = [["序号", "档号", "问题描述", "严重程度"]]
            idx = 1
            for result in validation_results:
                if not result.passed:
                    for issue in result.issues:
                        if issue.severity.value == "error":
                            issue_data.append([
                                str(idx),
                                result.archive_id,
                                issue.message,
                                "错误",
                            ])
                            idx += 1

            issue_table = Table(issue_data, colWidths=[1.5 * cm, 3.5 * cm, 8 * cm, 2 * cm])
            issue_table.setStyle(TableStyle([
                ("FONTNAME", (0, 0), (-1, -1), "ChineseFont" if self._has_chinese_font() else "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightcoral),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
            ]))
            story.append(issue_table)
            story.append(Spacer(1, 0.8 * cm))

        story.append(Paragraph("三、签章区域", heading_style))

        signature_data = [
            ["移交单位经办人:", "____________________", "接收单位经办人:", "____________________"],
            ["", "", "", ""],
            ["日期:", "______年____月____日", "日期:", "______年____月____日"],
            ["", "", "", ""],
            ["单位盖章:", "____________________", "单位盖章:", "____________________"],
        ]

        sig_table = Table(signature_data, colWidths=[3 * cm, 4 * cm, 3 * cm, 4 * cm])
        sig_table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), "ChineseFont" if self._has_chinese_font() else "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 11),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 15),
            ("TOPPADDING", (0, 0), (-1, -1), 15),
        ]))
        story.append(sig_table)

        doc.build(story)

        if self.logger:
            self.logger.info(
                f"生成接收回执: {output_path}",
                operation_type="generate_receipt",
                obj=output_path,
            )

        return output_path

    def _has_chinese_font(self):
        try:
            if SimpleDocTemplate:
                return "ChineseFont" in pdfmetrics.getRegisteredFontNames()
        except Exception:
            pass
        return False

    def generate_quality_report(self, scan_results: List[Any],
                                output_path: str = "质检报告.xlsx") -> str:
        if Workbook is None:
            raise ImportError("需要安装 openpyxl 库来生成质检报告")

        wb = Workbook()
        ws = wb.active
        ws.title = "质检汇总"

        header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
        normal_font = Font(name="微软雅黑", size=10)
        center_align = Alignment(horizontal="center", vertical="center")
        fill_header = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

        from .scanner import QualitySeverity

        total = len(scan_results)
        passed = sum(1 for r in scan_results if r.overall == QualitySeverity.PASS)
        warnings = sum(1 for r in scan_results if r.overall == QualitySeverity.WARNING)
        failed = sum(1 for r in scan_results if r.overall == QualitySeverity.FAIL)
        blank_pages = sum(1 for r in scan_results if r.blank_page)
        avg_dpi = sum(r.dpi for r in scan_results) / total if total > 0 else 0

        summary_items = [
            ["文件总数", str(total)],
            ["合格", str(passed)],
            ["警告", str(warnings)],
            ["不合格", str(failed)],
            ["合格率", f"{passed/total*100:.1f}%" if total > 0 else "0%"],
            ["平均分辨率", f"{avg_dpi:.0f} dpi"],
            ["空白页数", str(blank_pages)],
        ]

        for i, (label, value) in enumerate(summary_items):
            row = i + 1
            ws.cell(row=row, column=1, value=label).font = header_font
            ws.cell(row=row, column=2, value=value).font = normal_font

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 30

        ws2 = wb.create_sheet("详细结果")
        detail_headers = ["文件路径", "页码", "分辨率(dpi)", "色彩模式", "宽度(px)", "高度(px)",
                          "倾斜度(°)", "内容占比", "空白页", "综合结果"]
        for col, header in enumerate(detail_headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = center_align
            cell.fill = fill_header

        for i, result in enumerate(scan_results, 1):
            row = i + 1
            ws2.cell(row=row, column=1, value=result.file_path).font = normal_font
            ws2.cell(row=row, column=2, value=result.page).font = normal_font
            ws2.cell(row=row, column=3, value=f"{result.dpi:.0f}").font = normal_font
            ws2.cell(row=row, column=4, value=result.color_mode).font = normal_font
            ws2.cell(row=row, column=5, value=result.width).font = normal_font
            ws2.cell(row=row, column=6, value=result.height).font = normal_font
            ws2.cell(row=row, column=7, value=f"{result.tilt_degrees:.2f}").font = normal_font
            ws2.cell(row=row, column=8, value=f"{result.content_ratio*100:.1f}%").font = normal_font
            ws2.cell(row=row, column=9, value="是" if result.blank_page else "否").font = normal_font

            overall_cell = ws2.cell(row=row, column=10, value=result.overall.value)
            overall_cell.font = normal_font
            if result.overall == QualitySeverity.PASS:
                overall_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif result.overall == QualitySeverity.WARNING:
                overall_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            else:
                overall_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        widths = [40, 8, 12, 10, 10, 10, 10, 10, 8, 10]
        for i, width in enumerate(widths, 1):
            ws2.column_dimensions[get_column_letter(i)].width = width

        wb.save(output_path)

        if self.logger:
            self.logger.info(
                f"生成质检报告: {output_path}",
                operation_type="generate_quality_report",
                obj=output_path,
            )

        return output_path
