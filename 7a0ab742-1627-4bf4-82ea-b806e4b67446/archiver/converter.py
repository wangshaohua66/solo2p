import csv
import os
import json
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
from enum import Enum

try:
    from lxml import etree
except ImportError:
    etree = None

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    Workbook = None
    load_workbook = None


class InputFormat(str, Enum):
    XML_DAT46 = "xml_dat46"
    XML_CUSTOM = "xml_custom"
    EXCEL = "excel"
    CSV = "csv"
    UNKNOWN = "unknown"


class FormatConverter:
    STANDARD_FIELDS = [
        "archive_number", "title", "author", "created_date", "archived_date",
        "retention_period", "secrecy_level", "file_name", "file_size",
        "file_format", "page_count", "summary", "keywords", "category",
    ]

    def __init__(self, logger=None):
        self.logger = logger

    def detect_format(self, file_path: str) -> InputFormat:
        ext = Path(file_path).suffix.lower()

        if ext in [".xml"]:
            return self._detect_xml_format(file_path)
        elif ext in [".xlsx", ".xls"]:
            return InputFormat.EXCEL
        elif ext in [".csv"]:
            return InputFormat.CSV
        else:
            return InputFormat.UNKNOWN

    def _detect_xml_format(self, file_path: str) -> InputFormat:
        if etree is None:
            return InputFormat.XML_CUSTOM

        try:
            tree = etree.parse(file_path)
            root = tree.getroot()

            if "DA/T46" in root.tag or "电子档案" in root.tag:
                return InputFormat.XML_DAT46

            ns = root.nsmap
            if ns and any("dat46" in str(v).lower() for v in ns.values() if v):
                return InputFormat.XML_DAT46

            return InputFormat.XML_CUSTOM
        except Exception:
            return InputFormat.XML_CUSTOM

    def parse(self, file_path: str, format_type: Optional[InputFormat] = None) -> List[Dict[str, Any]]:
        if format_type is None:
            format_type = self.detect_format(file_path)

        if format_type == InputFormat.XML_DAT46:
            return self._parse_dat46_xml(file_path)
        elif format_type == InputFormat.XML_CUSTOM:
            return self._parse_custom_xml(file_path)
        elif format_type == InputFormat.EXCEL:
            return self._parse_excel(file_path)
        elif format_type == InputFormat.CSV:
            return self._parse_csv(file_path)
        else:
            raise ValueError(f"不支持的文件格式: {format_type}")

    def _parse_dat46_xml(self, file_path: str) -> List[Dict[str, Any]]:
        if etree is None:
            raise ImportError("需要安装 lxml 库来解析 XML 文件")

        tree = etree.parse(file_path)
        root = tree.getroot()
        archives = []

        ns = root.nsmap
        ns_uri = None
        if ns:
            for prefix, uri in ns.items():
                if uri:
                    ns_uri = uri
                    break

        def find_records(element):
            records = []
            tag_local = etree.QName(element.tag).localname
            if tag_local in ["档案", "record", "archive"] and len(element) > 1:
                records.append(element)
            for child in element:
                records.extend(find_records(child))
            return records

        record_elements = find_records(root)

        for record in record_elements:
            archive = {}
            for elem in record.iter():
                tag_local = etree.QName(elem.tag).localname
                if tag_local != etree.QName(record.tag).localname and len(elem) == 0:
                    mapped_key = self._map_field_name(tag_local)
                    if mapped_key:
                        archive[mapped_key] = elem.text or ""
                    else:
                        archive[tag_local] = elem.text or ""
            if archive:
                archives.append(archive)

        if not archives:
            archive = {}
            for elem in root.iter():
                tag_local = etree.QName(elem.tag).localname
                if len(elem) == 0 and elem.text:
                    mapped_key = self._map_field_name(tag_local)
                    if mapped_key:
                        archive[mapped_key] = elem.text or ""
                    else:
                        archive[tag_local] = elem.text or ""
            if archive:
                archives.append(archive)

        if self.logger:
            self.logger.info(
                f"解析 DA/T 46-2009 XML 文件: {file_path}, 共 {len(archives)} 条记录",
                operation_type="parse_xml",
                obj=file_path,
            )

        return archives

    def _parse_custom_xml(self, file_path: str) -> List[Dict[str, Any]]:
        if etree is None:
            raise ImportError("需要安装 lxml 库来解析 XML 文件")

        tree = etree.parse(file_path)
        root = tree.getroot()
        archives = []

        def find_record_elements(element):
            records = []
            if len(element) > 1:
                has_child_with_children = sum(1 for c in element if len(c) > 0)
                if has_child_with_children > 0 and has_child_with_children < len(element):
                    pass
                else:
                    records.append(element)
            for child in element:
                if len(child) > 1:
                    records.extend(find_record_elements(child))
            return records

        record_elements = find_record_elements(root)

        if not record_elements:
            record_elements = [root]

        for record in record_elements:
            archive = {}
            self._flatten_xml(record, archive, "")
            if archive:
                archives.append(archive)

        if self.logger:
            self.logger.info(
                f"解析自定义 XML 文件: {file_path}, 共 {len(archives)} 条记录",
                operation_type="parse_xml",
                obj=file_path,
            )

        return archives

    def _flatten_xml(self, element, result: Dict[str, Any], prefix: str):
        for child in element:
            tag = etree.QName(child.tag).localname if isinstance(child.tag, str) and "}" in child.tag else child.tag
            if prefix:
                key = f"{prefix}_{tag}"
            else:
                key = tag

            if len(child) == 0:
                mapped_key = self._map_field_name(key)
                result[mapped_key or key] = child.text or ""
            else:
                self._flatten_xml(child, result, key)

    def _map_field_name(self, tag: str) -> Optional[str]:
        field_mapping = {
            "档号": "archive_number",
            "题名": "title",
            "责任者": "author",
            "形成时间": "created_date",
            "归档时间": "archived_date",
            "保管期限": "retention_period",
            "密级": "secrecy_level",
            "文件名": "file_name",
            "文件大小": "file_size",
            "文件格式": "file_format",
            "页数": "page_count",
            "摘要": "summary",
            "主题词": "keywords",
            "类别": "category",
            "archiveNumber": "archive_number",
            "ArchiveNumber": "archive_number",
        }
        return field_mapping.get(tag)

    def _parse_excel(self, file_path: str) -> List[Dict[str, Any]]:
        if load_workbook is None:
            raise ImportError("需要安装 openpyxl 库来解析 Excel 文件")

        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        archives = []

        headers = []
        for cell in ws[1]:
            header = self._map_field_name(str(cell.value).strip()) if cell.value else None
            headers.append(header or (str(cell.value).strip() if cell.value else f"column_{cell.column}"))

        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None or v == "" for v in row):
                continue

            archive = {}
            for i, value in enumerate(row):
                if i < len(headers):
                    key = headers[i]
                    archive[key] = value if value is not None else ""
            if archive:
                archives.append(archive)

        if self.logger:
            self.logger.info(
                f"解析 Excel 文件: {file_path}, 共 {len(archives)} 条记录",
                operation_type="parse_excel",
                obj=file_path,
            )

        return archives

    def _parse_csv(self, file_path: str) -> List[Dict[str, Any]]:
        archives = []

        encodings = ["utf-8", "gbk", "gb2312", "utf-16"]
        for encoding in encodings:
            try:
                with open(file_path, "r", encoding=encoding) as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        archive = {}
                        for key, value in row.items():
                            mapped_key = self._map_field_name(key.strip())
                            archive[mapped_key or key.strip()] = value
                        if archive:
                            archives.append(archive)
                break
            except UnicodeDecodeError:
                continue

        if self.logger:
            self.logger.info(
                f"解析 CSV 文件: {file_path}, 共 {len(archives)} 条记录",
                operation_type="parse_csv",
                obj=file_path,
            )

        return archives

    def to_dat46_xml(self, archives: List[Dict[str, Any]], output_path: str):
        if etree is None:
            raise ImportError("需要安装 lxml 库来生成 XML 文件")

        root = etree.Element("电子档案移交清单")
        root.set("xmlns", "http://www.saac.gov.cn/dat46")
        root.set("版本", "DA/T 46-2009")

        archives_elem = etree.SubElement(root, "档案列表")

        for archive in archives:
            record = etree.SubElement(archives_elem, "档案")

            field_mapping = {
                "archive_number": "档号",
                "title": "题名",
                "author": "责任者",
                "created_date": "形成时间",
                "archived_date": "归档时间",
                "retention_period": "保管期限",
                "secrecy_level": "密级",
                "file_name": "文件名",
                "file_size": "文件大小",
                "file_format": "文件格式",
                "page_count": "页数",
                "summary": "摘要",
                "keywords": "主题词",
                "category": "类别",
            }

            for key, value in archive.items():
                tag_name = field_mapping.get(key, key)
                elem = etree.SubElement(record, tag_name)
                elem.text = str(value) if value is not None else ""

        tree = etree.ElementTree(root)
        tree.write(output_path, pretty_print=True, encoding="UTF-8", xml_declaration=True)

        if self.logger:
            self.logger.info(
                f"生成 DA/T 46-2009 XML 文件: {output_path}, 共 {len(archives)} 条记录",
                operation_type="generate_xml",
                obj=output_path,
            )

    def to_excel(self, archives: List[Dict[str, Any]], output_path: str, sheet_name: str = "档案清单"):
        if Workbook is None:
            raise ImportError("需要安装 openpyxl 库来生成 Excel 文件")

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        all_keys = self.STANDARD_FIELDS.copy()
        for archive in archives:
            for key in archive.keys():
                if key not in all_keys:
                    all_keys.append(key)

        header_mapping = {
            "archive_number": "档号",
            "title": "题名",
            "author": "责任者",
            "created_date": "形成时间",
            "archived_date": "归档时间",
            "retention_period": "保管期限",
            "secrecy_level": "密级",
            "file_name": "文件名",
            "file_size": "文件大小",
            "file_format": "文件格式",
            "page_count": "页数",
            "summary": "摘要",
            "keywords": "主题词",
            "category": "类别",
        }

        headers = [header_mapping.get(k, k) for k in all_keys]
        ws.append(headers)

        for archive in archives:
            row = [archive.get(key, "") for key in all_keys]
            ws.append(row)

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(output_path)

        if self.logger:
            self.logger.info(
                f"生成 Excel 文件: {output_path}, 共 {len(archives)} 条记录",
                operation_type="generate_excel",
                obj=output_path,
            )

    def to_csv(self, archives: List[Dict[str, Any]], output_path: str):
        all_keys = self.STANDARD_FIELDS.copy()
        for archive in archives:
            for key in archive.keys():
                if key not in all_keys:
                    all_keys.append(key)

        with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=all_keys)
            writer.writeheader()
            for archive in archives:
                writer.writerow(archive)

        if self.logger:
            self.logger.info(
                f"生成 CSV 文件: {output_path}, 共 {len(archives)} 条记录",
                operation_type="generate_csv",
                obj=output_path,
            )

    def to_json(self, archives: List[Dict[str, Any]], output_path: str):
        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(archives, f, ensure_ascii=False, indent=2)

        if self.logger:
            self.logger.info(
                f"生成 JSON 文件: {output_path}, 共 {len(archives)} 条记录",
                operation_type="generate_json",
                obj=output_path,
            )

    def convert(self, input_path: str, output_path: str, output_format: str = "xml"):
        archives = self.parse(input_path)

        output_format = output_format.lower()
        if output_format in ["xml", "dat46"]:
            self.to_dat46_xml(archives, output_path)
        elif output_format in ["excel", "xlsx"]:
            self.to_excel(archives, output_path)
        elif output_format == "csv":
            self.to_csv(archives, output_path)
        elif output_format == "json":
            self.to_json(archives, output_path)
        else:
            raise ValueError(f"不支持的输出格式: {output_format}")

        return archives

    def create_transfer_package(self, archives: List[Dict[str, Any]], output_dir: str,
                               file_dir: Optional[str] = None,
                               validation_results: Optional[List] = None):
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)

        metadata_dir = output_path / "元数据"
        files_dir = output_path / "电子文件"
        report_dir = output_path / "校验报告"

        metadata_dir.mkdir(exist_ok=True)
        files_dir.mkdir(exist_ok=True)
        report_dir.mkdir(exist_ok=True)

        valid_archives = []
        if validation_results:
            for i, result in enumerate(validation_results):
                if result.passed and i < len(archives):
                    valid_archives.append(archives[i])
        else:
            valid_archives = archives

        xml_path = metadata_dir / "电子档案元数据.xml"
        self.to_dat46_xml(valid_archives, str(xml_path))

        excel_path = metadata_dir / "电子档案清单.xlsx"
        self.to_excel(valid_archives, str(excel_path))

        if file_dir:
            import shutil
            src_dir = Path(file_dir)
            for archive in valid_archives:
                file_name = archive.get("file_name", "")
                if file_name and (src_dir / file_name).exists():
                    shutil.copy2(src_dir / file_name, files_dir / file_name)

        if validation_results:
            report_path = report_dir / "校验报告.json"
            with open(report_path, "w", encoding="utf-8") as f:
                json.dump(
                    [r.to_dict() for r in validation_results],
                    f, ensure_ascii=False, indent=2
                )

        if self.logger:
            self.logger.info(
                f"创建移交包: {output_dir}, 有效档案 {len(valid_archives)} 件",
                operation_type="create_package",
                obj=str(output_dir),
            )

        return str(output_path)
