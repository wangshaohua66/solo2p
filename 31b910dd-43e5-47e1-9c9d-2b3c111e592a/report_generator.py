import csv
import json
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from config import REPORT_DIR, REVIEW_MODULES, ISSUE_SEVERITY, DRUG_TYPE_CONFIG
from logger import logger

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    )
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    PDF_REPORT_AVAILABLE = True
except ImportError:
    PDF_REPORT_AVAILABLE = False
    logger.warning("reportlab未安装，将降级生成文本格式报告")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logger.warning("openpyxl未安装，将降级生成CSV格式问题表")


SEVERITY_LABELS = {
    "FATAL": "致命错误",
    "DEFECT": "一般缺陷",
    "SUGGESTION": "建议优化",
}

RECOMMENDATION_LABELS = {
    "REJECT": "不予受理（存在致命错误）",
    "REVISE": "退回修改",
    "SUPPLEMENT": "补充资料",
    "PASS": "形式审查通过",
}

MODULE_LABELS = {
    "ctd_structure": "CTD目录结构",
    "file_naming": "文件命名规范",
    "page_continuity": "页码连续性",
    "signature_seal": "签字盖章识别",
    "overview_check": "综述内容抽检",
    "cross_validate": "交叉校验",
}


class ReportGenerator:
    def __init__(self, project_name: str, output_dir: Optional[Path] = None) -> None:
        self.project_name = project_name
        self.output_dir = output_dir or REPORT_DIR
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.safe_name = "".join(c for c in project_name if c.isalnum() or c in "-_ ")
        self.safe_name = self.safe_name.strip().replace(" ", "_")

    def _register_fonts(self) -> None:
        if not PDF_REPORT_AVAILABLE:
            return
        font_paths = [
            "/System/Library/Fonts/PingFang.ttc",
            "/System/Library/Fonts/STHeiti Medium.ttc",
            "/Library/Fonts/Arial Unicode.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        ]
        for font_path in font_paths:
            if Path(font_path).exists():
                try:
                    pdfmetrics.registerFont(TTFont("ChineseFont", font_path))
                    return
                except Exception:
                    continue
        logger.warning("未找到中文字体，PDF报告中文可能显示异常")

    def generate_reports(
        self,
        issues: List[Any],
        stats: Dict[str, Any],
        project_info: Dict[str, Any],
        structure_diff: Optional[List[str]] = None,
    ) -> Dict[str, Path]:
        logger.info("开始生成审查报告")
        start_time = time.time()
        result_paths: Dict[str, Path] = {}

        pdf_path = self._generate_pdf_report(issues, stats, project_info, structure_diff)
        if pdf_path:
            result_paths["pdf"] = pdf_path

        excel_path = self._generate_excel_report(issues, stats, project_info)
        if excel_path:
            result_paths["excel"] = excel_path

        json_path = self._generate_json_report(issues, stats, project_info, structure_diff)
        result_paths["json"] = json_path

        txt_path = self._generate_text_summary(issues, stats, project_info, structure_diff)
        result_paths["text"] = txt_path

        logger.info(f"报告生成完成，耗时 {time.time() - start_time:.2f} 秒")
        for fmt, path in result_paths.items():
            logger.info(f"  {fmt.upper()}: {path}")
        return result_paths

    def _generate_pdf_report(
        self,
        issues: List[Any],
        stats: Dict[str, Any],
        project_info: Dict[str, Any],
        structure_diff: Optional[List[str]],
    ) -> Optional[Path]:
        if not PDF_REPORT_AVAILABLE:
            return None

        try:
            self._register_fonts()
            pdf_path = self.output_dir / f"{self.safe_name}_审查报告_{self.timestamp}.pdf"
            doc = SimpleDocTemplate(
                str(pdf_path),
                pagesize=A4,
                leftMargin=2 * cm,
                rightMargin=2 * cm,
                topMargin=2.5 * cm,
                bottomMargin=2.5 * cm,
            )

            styles = getSampleStyleSheet()
            font_name = "ChineseFont" if "ChineseFont" in pdfmetrics.getRegisteredFontNames() else "Helvetica"

            title_style = ParagraphStyle(
                "CustomTitle", parent=styles["Title"],
                fontName=font_name, fontSize=20, leading=28, spaceAfter=20,
            )
            h1_style = ParagraphStyle(
                "CustomH1", parent=styles["Heading1"],
                fontName=font_name, fontSize=16, leading=22, spaceBefore=15, spaceAfter=10,
                textColor=colors.HexColor("#1f4e79"),
            )
            h2_style = ParagraphStyle(
                "CustomH2", parent=styles["Heading2"],
                fontName=font_name, fontSize=13, leading=18, spaceBefore=10, spaceAfter=6,
                textColor=colors.HexColor("#2e75b6"),
            )
            body_style = ParagraphStyle(
                "CustomBody", parent=styles["Normal"],
                fontName=font_name, fontSize=10, leading=16,
            )
            small_style = ParagraphStyle(
                "CustomSmall", parent=styles["Normal"],
                fontName=font_name, fontSize=9, leading=14, textColor=colors.grey,
            )

            story = []
            story.append(Paragraph(f"药品注册申报资料形式审查报告", title_style))
            story.append(Paragraph(f"项目名称: {self.project_name}", body_style))
            story.append(Paragraph(
                f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", body_style
            ))
            story.append(Spacer(1, 12))

            drug_label = DRUG_TYPE_CONFIG.get(
                project_info.get("drug_type", "chemical"), {}
            ).get("label", "未知")
            summary_data = [
                ["项目", "内容"],
                ["药品类型", drug_label],
                ["申请人", project_info.get("applicant", "—")],
                ["资料路径", str(project_info.get("folder_path", "—"))],
                ["文件总数", str(project_info.get("total_files", 0))],
                ["资料大小", f"{project_info.get('total_size_mb', 0):.2f} MB"],
                ["综合评分", f"{stats.get('overall_score', 0)} / 100"],
                ["审查建议", RECOMMENDATION_LABELS.get(stats.get("recommendation", ""), "—")],
            ]
            summary_table = Table(summary_data, colWidths=[4 * cm, 11 * cm])
            summary_table.setStyle(TableStyle([
                ("FONT", (0, 0), (-1, -1), font_name, 10),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f4e79")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("BACKGROUND", (0, 1), (0, -1), colors.HexColor("#d6e4f0")),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]))
            story.append(summary_table)
            story.append(Spacer(1, 15))

            story.append(Paragraph("问题统计", h1_style))
            severity_data = [
                ["严重程度", "问题数量", "说明"],
                [SEVERITY_LABELS.get("FATAL", "致命错误"), str(stats["by_severity"].get("FATAL", 0)),
                 "违反强制性要求，直接退回"],
                [SEVERITY_LABELS.get("DEFECT", "一般缺陷"), str(stats["by_severity"].get("DEFECT", 0)),
                 "不符合规范，需补充或修改"],
                [SEVERITY_LABELS.get("SUGGESTION", "建议优化"), str(stats["by_severity"].get("SUGGESTION", 0)),
                 "规范性优化建议"],
            ]
            sev_table = Table(severity_data, colWidths=[3.5 * cm, 2.5 * cm, 9 * cm])
            sev_table.setStyle(TableStyle([
                ("FONT", (0, 0), (-1, -1), font_name, 10),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f4e79")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#ffd7d7")),
                ("BACKGROUND", (0, 2), (-1, 2), colors.HexColor("#fff3cd")),
                ("BACKGROUND", (0, 3), (-1, 3), colors.HexColor("#d1ecf1")),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]))
            story.append(sev_table)
            story.append(Spacer(1, 15))

            story.append(Paragraph("各模块得分", h1_style))
            module_data = [["检查模块", "权重", "得分", "问题数"]]
            for mod in REVIEW_MODULES:
                key = mod["key"]
                score = stats["module_scores"].get(key, 0)
                issue_count = stats["by_module"].get(key, 0)
                module_data.append([
                    MODULE_LABELS.get(key, key),
                    f"{int(mod['weight'] * 100)}%",
                    str(score),
                    str(issue_count),
                ])
            mod_table = Table(module_data, colWidths=[5 * cm, 2.5 * cm, 2.5 * cm, 2.5 * cm])
            mod_table.setStyle(TableStyle([
                ("FONT", (0, 0), (-1, -1), font_name, 10),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f4e79")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (1, 0), (-1, -1), "CENTER"),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
                ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]))
            story.append(mod_table)
            story.append(Spacer(1, 15))

            story.append(PageBreak())
            story.append(Paragraph("问题清单", h1_style))

            for mod in REVIEW_MODULES:
                mod_key = mod["key"]
                mod_issues = [i for i in issues if getattr(i, "module", "") == mod_key]
                if not mod_issues:
                    continue
                story.append(Paragraph(f"{MODULE_LABELS.get(mod_key, mod_key)}", h2_style))

                issue_header = ["序号", "严重程度", "问题类型", "问题描述", "文件路径", "整改建议"]
                issue_rows = [issue_header]
                for idx, issue in enumerate(mod_issues, 1):
                    issue_rows.append([
                        str(idx),
                        SEVERITY_LABELS.get(getattr(issue, "severity", ""), ""),
                        getattr(issue, "issue_type", ""),
                        getattr(issue, "description", "")[:150],
                        Path(getattr(issue, "file_path", "")).name[:50],
                        getattr(issue, "suggestion", "")[:100],
                    ])
                if len(issue_rows) > 1:
                    issue_table = Table(
                        issue_rows,
                        colWidths=[1 * cm, 2 * cm, 2.5 * cm, 5 * cm, 2.5 * cm, 3 * cm],
                        repeatRows=1,
                    )
                    issue_table.setStyle(TableStyle([
                        ("FONT", (0, 0), (-1, -1), font_name, 8),
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2e75b6")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("GRID", (0, 0), (-1, -1), 0.3, colors.lightgrey),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("LEFTPADDING", (0, 0), (-1, -1), 4),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                        ("TOPPADDING", (0, 0), (-1, -1), 3),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                    ]))
                    story.append(issue_table)
                    story.append(Spacer(1, 8))

            if structure_diff:
                story.append(PageBreak())
                story.append(Paragraph("CTD目录结构差异", h1_style))
                for line in structure_diff:
                    clean_line = line.replace("\033[91m", "").replace("\033[92m", "")
                    clean_line = clean_line.replace("\033[93m", "").replace("\033[0m", "")
                    story.append(Paragraph(clean_line, small_style))

            doc.build(story)
            logger.info(f"PDF报告已生成: {pdf_path}")
            return pdf_path
        except Exception as e:
            logger.error(f"PDF报告生成失败: {e}", exception=e)
            return None

    def _generate_excel_report(
        self,
        issues: List[Any],
        stats: Dict[str, Any],
        project_info: Dict[str, Any],
    ) -> Optional[Path]:
        if not EXCEL_AVAILABLE:
            return self._generate_csv_report(issues, stats, project_info)

        try:
            xlsx_path = self.output_dir / f"{self.safe_name}_问题汇总_{self.timestamp}.xlsx"
            wb = openpyxl.Workbook()

            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
            thin_border = Border(
                left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"),
            )
            fatal_fill = PatternFill(start_color="FFD7D7", end_color="FFD7D7", fill_type="solid")
            defect_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
            suggest_fill = PatternFill(start_color="D1ECF1", end_color="D1ECF1", fill_type="solid")

            ws_summary = wb.active
            ws_summary.title = "审查概要"
            summary_items = [
                ("项目名称", self.project_name),
                ("药品类型", DRUG_TYPE_CONFIG.get(project_info.get("drug_type", "chemical"), {}).get("label", "")),
                ("申请人", project_info.get("applicant", "")),
                ("资料路径", str(project_info.get("folder_path", ""))),
                ("文件总数", project_info.get("total_files", 0)),
                ("资料大小(MB)", round(project_info.get("total_size_mb", 0), 2)),
                ("综合评分", stats.get("overall_score", 0)),
                ("审查建议", RECOMMENDATION_LABELS.get(stats.get("recommendation", ""), "")),
                ("致命错误数", stats["by_severity"].get("FATAL", 0)),
                ("一般缺陷数", stats["by_severity"].get("DEFECT", 0)),
                ("建议优化数", stats["by_severity"].get("SUGGESTION", 0)),
                ("常见问题数", stats.get("common_count", 0)),
                ("问题总数", stats.get("total", 0)),
                ("生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ]
            for r, (key, val) in enumerate(summary_items, 1):
                ws_summary.cell(row=r, column=1, value=key).font = Font(bold=True)
                ws_summary.cell(row=r, column=2, value=str(val))
                for c in (1, 2):
                    cell = ws_summary.cell(row=r, column=c)
                    cell.border = thin_border
                    cell.alignment = left_align
            ws_summary.column_dimensions["A"].width = 20
            ws_summary.column_dimensions["B"].width = 60

            ws_modules = wb.create_sheet("模块得分")
            mod_headers = ["检查模块", "权重", "得分", "问题数量"]
            for c, h in enumerate(mod_headers, 1):
                cell = ws_modules.cell(row=1, column=c, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border
            for r, mod in enumerate(REVIEW_MODULES, 2):
                key = mod["key"]
                row_data = [
                    MODULE_LABELS.get(key, key),
                    f"{int(mod['weight'] * 100)}%",
                    stats["module_scores"].get(key, 0),
                    stats["by_module"].get(key, 0),
                ]
                for c, val in enumerate(row_data, 1):
                    cell = ws_modules.cell(row=r, column=c, value=val)
                    cell.border = thin_border
                    cell.alignment = center_align if c > 1 else left_align
            ws_modules.column_dimensions["A"].width = 25
            for col in ["B", "C", "D"]:
                ws_modules.column_dimensions[col].width = 12

            ws_issues = wb.create_sheet("问题清单")
            issue_headers = [
                "序号", "严重程度", "模块", "问题类型", "问题描述",
                "文件路径", "整改建议", "是否常见问题", "历史出现次数",
            ]
            for c, h in enumerate(issue_headers, 1):
                cell = ws_issues.cell(row=1, column=c, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border

            fill_map = {"FATAL": fatal_fill, "DEFECT": defect_fill, "SUGGESTION": suggest_fill}
            for r, issue in enumerate(issues, 2):
                severity = getattr(issue, "severity", "")
                row_data = [
                    r - 1,
                    SEVERITY_LABELS.get(severity, severity),
                    MODULE_LABELS.get(getattr(issue, "module", ""), getattr(issue, "module", "")),
                    getattr(issue, "issue_type", ""),
                    getattr(issue, "description", ""),
                    getattr(issue, "file_path", ""),
                    getattr(issue, "suggestion", ""),
                    "是" if getattr(issue, "is_common", False) else "否",
                    getattr(issue, "occurrence_count", 0),
                ]
                row_fill = fill_map.get(severity)
                for c, val in enumerate(row_data, 1):
                    cell = ws_issues.cell(row=r, column=c, value=val)
                    cell.border = thin_border
                    cell.alignment = left_align if c in (4, 5, 6, 7) else center_align
                    if row_fill:
                        cell.fill = row_fill

            col_widths = [6, 12, 18, 18, 50, 40, 40, 12, 14]
            for c, w in enumerate(col_widths, 1):
                ws_issues.column_dimensions[get_column_letter(c)].width = w
            ws_issues.freeze_panes = "A2"

            ws_stats = wb.create_sheet("问题统计")
            type_headers = ["问题类型", "类别", "严重程度", "出现次数", "说明"]
            for c, h in enumerate(type_headers, 1):
                cell = ws_stats.cell(row=1, column=c, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border

            from issue_classifier import ISSUE_TYPE_MAPPING
            r = 2
            for issue_type, count in stats["by_type"].most_common():
                mapping = ISSUE_TYPE_MAPPING.get(issue_type, {})
                row_data = [
                    issue_type,
                    mapping.get("category", ""),
                    SEVERITY_LABELS.get(mapping.get("severity", ""), mapping.get("severity", "")),
                    count,
                    mapping.get("description", ""),
                ]
                for c, val in enumerate(row_data, 1):
                    cell = ws_stats.cell(row=r, column=c, value=val)
                    cell.border = thin_border
                    cell.alignment = center_align if c != 5 else left_align
                r += 1
            ws_stats.column_dimensions["A"].width = 28
            ws_stats.column_dimensions["B"].width = 14
            ws_stats.column_dimensions["C"].width = 12
            ws_stats.column_dimensions["D"].width = 10
            ws_stats.column_dimensions["E"].width = 30

            wb.save(str(xlsx_path))
            logger.info(f"Excel报告已生成: {xlsx_path}")
            return xlsx_path
        except Exception as e:
            logger.error(f"Excel报告生成失败: {e}", exception=e)
            return self._generate_csv_report(issues, stats, project_info)

    def _generate_csv_report(
        self, issues: List[Any], stats: Dict[str, Any], project_info: Dict[str, Any]
    ) -> Optional[Path]:
        try:
            csv_path = self.output_dir / f"{self.safe_name}_问题汇总_{self.timestamp}.csv"
            with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                writer.writerow([
                    "序号", "严重程度", "模块", "问题类型",
                    "问题描述", "文件路径", "整改建议", "是否常见问题",
                ])
                for r, issue in enumerate(issues, 1):
                    writer.writerow([
                        r,
                        SEVERITY_LABELS.get(getattr(issue, "severity", ""), ""),
                        MODULE_LABELS.get(getattr(issue, "module", ""), ""),
                        getattr(issue, "issue_type", ""),
                        getattr(issue, "description", ""),
                        getattr(issue, "file_path", ""),
                        getattr(issue, "suggestion", ""),
                        "是" if getattr(issue, "is_common", False) else "否",
                    ])
            logger.info(f"CSV报告已生成: {csv_path}")
            return csv_path
        except Exception as e:
            logger.error(f"CSV报告生成失败: {e}", exception=e)
            return None

    def _generate_json_report(
        self,
        issues: List[Any],
        stats: Dict[str, Any],
        project_info: Dict[str, Any],
        structure_diff: Optional[List[str]],
    ) -> Path:
        json_path = self.output_dir / f"{self.safe_name}_审查数据_{self.timestamp}.json"
        issues_json = []
        for issue in issues:
            issues_json.append({
                "severity": getattr(issue, "severity", ""),
                "module": getattr(issue, "module", ""),
                "issue_type": getattr(issue, "issue_type", ""),
                "description": getattr(issue, "description", ""),
                "file_path": getattr(issue, "file_path", ""),
                "suggestion": getattr(issue, "suggestion", ""),
                "is_common": getattr(issue, "is_common", False),
                "occurrence_count": getattr(issue, "occurrence_count", 0),
            })

        def convert_counter(counter_like):
            if hasattr(counter_like, "items"):
                return dict(counter_like)
            return counter_like

        report_data = {
            "project_name": self.project_name,
            "generated_at": datetime.now().isoformat(),
            "project_info": {k: (str(v) if isinstance(v, Path) else v) for k, v in project_info.items()},
            "statistics": {
                "total": stats.get("total", 0),
                "overall_score": stats.get("overall_score", 0),
                "recommendation": stats.get("recommendation", ""),
                "common_count": stats.get("common_count", 0),
                "by_severity": convert_counter(stats.get("by_severity", {})),
                "by_module": convert_counter(stats.get("by_module", {})),
                "by_type": convert_counter(stats.get("by_type", {})),
                "module_scores": stats.get("module_scores", {}),
            },
            "issues": issues_json,
            "structure_diff": [
                l.replace("\033[91m", "").replace("\033[92m", "").replace("\033[93m", "").replace("\033[0m", "")
                for l in (structure_diff or [])
            ],
        }
        json_path.write_text(
            json.dumps(report_data, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        logger.info(f"JSON数据已导出: {json_path}")
        return json_path

    def _generate_text_summary(
        self,
        issues: List[Any],
        stats: Dict[str, Any],
        project_info: Dict[str, Any],
        structure_diff: Optional[List[str]],
    ) -> Path:
        txt_path = self.output_dir / f"{self.safe_name}_审查摘要_{self.timestamp}.txt"
        lines = []
        lines.append("=" * 70)
        lines.append("药品注册申报资料形式审查报告")
        lines.append("=" * 70)
        lines.append(f"项目名称: {self.project_name}")
        lines.append(f"药品类型: {DRUG_TYPE_CONFIG.get(project_info.get('drug_type', 'chemical'), {}).get('label', '')}")
        lines.append(f"申请人: {project_info.get('applicant', '—')}")
        lines.append(f"资料路径: {project_info.get('folder_path', '—')}")
        lines.append(f"文件总数: {project_info.get('total_files', 0)}")
        lines.append(f"资料大小: {project_info.get('total_size_mb', 0):.2f} MB")
        lines.append(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append("")
        lines.append("-" * 70)
        lines.append("审查结论")
        lines.append("-" * 70)
        lines.append(f"综合评分: {stats.get('overall_score', 0)} / 100")
        lines.append(f"审查建议: {RECOMMENDATION_LABELS.get(stats.get('recommendation', ''), '')}")
        lines.append(f"问题总数: {stats.get('total', 0)}")
        lines.append(f"  致命错误: {stats['by_severity'].get('FATAL', 0)}")
        lines.append(f"  一般缺陷: {stats['by_severity'].get('DEFECT', 0)}")
        lines.append(f"  建议优化: {stats['by_severity'].get('SUGGESTION', 0)}")
        lines.append(f"常见问题: {stats.get('common_count', 0)} 个")
        lines.append("")
        lines.append("-" * 70)
        lines.append("各模块得分")
        lines.append("-" * 70)
        for mod in REVIEW_MODULES:
            key = mod["key"]
            lines.append(
                f"  {MODULE_LABELS.get(key, key):<15} "
                f"得分 {stats['module_scores'].get(key, 0):>3} | "
                f"问题 {stats['by_module'].get(key, 0):>3} 个"
            )
        lines.append("")
        lines.append("-" * 70)
        lines.append("问题清单（按严重程度）")
        lines.append("-" * 70)
        for sev in ("FATAL", "DEFECT", "SUGGESTION"):
            sev_issues = [i for i in issues if getattr(i, "severity", "") == sev]
            if not sev_issues:
                continue
            lines.append(f"\n[{SEVERITY_LABELS.get(sev, sev)}] 共 {len(sev_issues)} 项:")
            for idx, issue in enumerate(sev_issues, 1):
                lines.append(f"  {idx}. {MODULE_LABELS.get(getattr(issue, 'module', ''), '')} - "
                             f"{getattr(issue, 'description', '')[:80]}")
                if getattr(issue, "file_path", ""):
                    lines.append(f"     文件: {Path(getattr(issue, 'file_path', '')).name}")
                if getattr(issue, "suggestion", ""):
                    lines.append(f"     建议: {getattr(issue, 'suggestion', '')[:80]}")
        if structure_diff:
            lines.append("")
            lines.append("-" * 70)
            lines.append("CTD目录结构差异")
            lines.append("-" * 70)
            for line in structure_diff:
                clean = line.replace("\033[91m", "").replace("\033[92m", "").replace("\033[93m", "").replace("\033[0m", "")
                lines.append(clean)

        txt_path.write_text("\n".join(lines), encoding="utf-8")
        logger.info(f"文本摘要已生成: {txt_path}")
        return txt_path
