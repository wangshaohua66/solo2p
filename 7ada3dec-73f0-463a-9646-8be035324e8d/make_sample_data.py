"""
make_sample_data.py — 生成三种格式的临床试验测试数据

输出至 sample/ 目录：
  C01_rave.xml     —— Medidata Rave ODM 嵌套结构（中心 C01）
  C04_excel.xlsx   —— 多 Sheet 模板（中心 C04）
  C08_custom.csv   —— 自定义单表 CSV（中心 C08，含中文表头）

数据中故意混入：跨中心单位差异（GLUC: mg/dL vs mmol/L）、离群值、
缺失字段、日期倒序等，以触发校验与一致性分析。
运行: python3 make_sample_data.py [--large N]  生成大规模性能数据
"""
from __future__ import annotations

import csv
import random
import sys
from datetime import date, timedelta
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import Workbook

OUT = Path("sample")
OUT.mkdir(exist_ok=True)
random.seed(42)

STUDYID = "STUDY01"
DOMAINS = ["DM", "LB", "VS", "AE", "EX"]
LAB_TESTS = ["GLUC", "CREAT", "ALT", "AST", "HGB", "WBC", "PLT", "SODIUM"]
VS_TESTS = ["TEMPERATURE", "HEIGHT", "WEIGHT", "BP", "HR"]
LAB_REF = {
    "GLUC": ("mg/dL", (70, 110)), "CREAT": ("mg/dL", (0.6, 1.2)),
    "ALT": ("U/L", (5, 40)), "AST": ("U/L", (8, 40)),
    "HGB": ("g/dL", (12, 16)), "WBC": ("10^9/L", (3.5, 9.5)),
    "PLT": ("10^9/L", (125, 350)), "SODIUM": ("mmol/L", (135, 145)),
}


def daterange(start: date, days: int) -> date:
    return start + timedelta(days=random.randint(0, max(days, 0)))


def make_subjects(center: str, n: int, start_subj: int):
    rows = []
    base = date(2024, 1, 1)
    for i in range(n):
        sid = f"{center}-{start_subj + i:04d}"
        sex = random.choice(["M", "F", "X"])  # X 为非法枚举(触发校验)
        age = random.randint(18, 80)
        # 个别缺失年龄
        if i % 17 == 0:
            age = None
        rfic = base + timedelta(days=random.randint(0, 200))
        rows.append({
            "STUDYID": STUDYID, "SUBJID": sid, "SITEID": center,
            "SEX": sex, "AGE": age, "AGEU": "YEARS" if age else "",
            "RACE": "ASIAN", "ARM": random.choice(["A", "B", "PLACEBO", "Z"]),
            "COUNTRY": "CHN", "RFICDTC": rfic.isoformat(),
        })
    return rows


def make_lab(subjects, center: str, unit_map: dict, visits=2):
    rows = []
    for s in subjects:
        for v in range(1, visits + 1):
            for code in LAB_TESTS:
                default_unit, (lo, hi) = LAB_REF[code]
                unit = unit_map.get(code, default_unit)
                val = round(random.uniform(lo, hi), 1)
                # 注入离群值与缺失
                if random.random() < 0.05:
                    val = round(random.uniform(hi * 1.5, hi * 2.5), 1)
                if random.random() < 0.03:
                    val = None
                rows.append({
                    "STUDYID": STUDYID, "SUBJID": s["SUBJID"], "SITEID": center,
                    "VISITNUM": v, "VISIT": f"VISIT{v}",
                    "LBTESTCD": code, "LBTEST": code.title(),
                    "LBORRES": "" if val is None else str(val),
                    "LBORRESU": "" if val is None else unit,
                    "LBDTC": s["RFICDTC"],
                })
    return rows


def make_vs(subjects, center: str):
    rows = []
    for s in subjects:
        for v in range(1, 3):
            for code in VS_TESTS:
                val = {
                    "TEMPERATURE": round(random.uniform(36, 37.5), 1),
                    "HEIGHT": random.randint(150, 185),
                    "WEIGHT": random.randint(45, 85),
                    "BP": random.randint(100, 140),
                    "HR": random.randint(60, 100),
                }[code]
                rows.append({
                    "STUDYID": STUDYID, "SUBJID": s["SUBJID"], "SITEID": center,
                    "VISITNUM": v, "VSTESTCD": code, "VSTEST": code.title(),
                    "VSORRES": str(val), "VSORRESU": {
                        "TEMPERATURE": "C", "HEIGHT": "cm", "WEIGHT": "kg",
                        "BP": "mmHg", "HR": "bpm"}[code],
                    "VSDTC": s["RFICDTC"],
                })
    return rows


def make_ae(subjects, center: str):
    rows = []
    terms = ["恶心", "头痛", "皮疹", "腹泻", "乏力"]
    for s in subjects:
        if random.random() < 0.5:
            start = date.fromisoformat(s["RFICDTC"]) + timedelta(days=10)
            end = start - timedelta(days=1)  # 故意倒序触发校验
            rows.append({
                "STUDYID": STUDYID, "SUBJID": s["SUBJID"], "SITEID": center,
                "AETERM": random.choice(terms), "AESEV": random.choice(
                    ["MILD", "MODERATE", "SEVERE", "XXX"]),
                "AESER": random.choice(["Y", "N", "?"]),
                "AESTDTC": start.isoformat(), "AEENDTC": end.isoformat(),
            })
    return rows


def make_ex(subjects, center: str):
    rows = []
    for s in subjects:
        start = date.fromisoformat(s["RFICDTC"]) + timedelta(days=1)
        end = start + timedelta(days=28)
        rows.append({
            "STUDYID": STUDYID, "SUBJID": s["SUBJID"], "SITEID": center,
            "EXTRT": "TestDrug", "EXDOSE": random.choice([100, 200, -5]),
            "EXDOSU": "mg", "EXSTDTC": start.isoformat(), "EXENDTC": end.isoformat(),
        })
    return rows


# ----------------------------- XML (Rave) -----------------------------
def write_xml(center: str, subj_n: int):
    subjects = make_subjects(center, subj_n, 1)
    root = ET.Element("ClinicalData", StudyOID=STUDYID)
    for s in subjects:
        sd = ET.SubElement(root, "SubjectData", SubjectKey=s["SUBJID"],
                            StudyOID=STUDYID, SiteOID=center)
        # DM
        fd = ET.SubElement(sd, "FormData", FormOID="DM")
        ig = ET.SubElement(fd, "ItemGroupData", ItemGroupOID="IG_DM")
        for k, v in s.items():
            ET.SubElement(ig, "ItemData", ItemOID=k, Value="" if v is None else str(v))
        # LB
        labs = make_lab(subjects, center, {"GLUC": "mg/dL"})
        fd = ET.SubElement(sd, "FormData", FormOID="LB")
        for lb in labs:
            if lb["SUBJID"] != s["SUBJID"]:
                continue
            ig = ET.SubElement(fd, "ItemGroupData", ItemGroupOID="IG_LB")
            for k, v in lb.items():
                ET.SubElement(ig, "ItemData", ItemOID=k, Value=str(v))
    ET.ElementTree(root).write(OUT / f"{center}_rave.xml",
                               encoding="UTF-8", xml_declaration=True)


# ----------------------------- Excel -----------------------------
def write_excel(center: str, subj_n: int):
    subjects = make_subjects(center, subj_n, 200)
    wb = Workbook()
    # DM sheet (中文表头测试别名映射)
    ws = wb.active
    ws.title = "DM"
    headers = ["研究编号", "受试者编号", "中心编号", "性别", "年龄", "年龄单位",
               "种族", "分组", "国家", "知情同意日期"]
    ws.append(headers)
    for s in subjects:
        ws.append([s["STUDYID"], s["SUBJID"], s["SITEID"], s["SEX"], s["AGE"] or "",
                   s["AGEU"], s["RACE"], s["ARM"], s["COUNTRY"], s["RFICDTC"]])
    # LB sheet (GLUC 用 mmol/L 与 XML 的 mg/dL 形成跨中心单位差异)
    ws = wb.create_sheet("LB")
    ws.append(["研究编号", "受试者编号", "中心编号", "访视序号", "访视名称",
               "检验代码", "检验名称", "检验结果", "检验单位", "检验日期"])
    for lb in make_lab(subjects, center, {"GLUC": "mmol/L"}):
        ws.append([lb["STUDYID"], lb["SUBJID"], lb["SITEID"], lb["VISITNUM"],
                   lb["VISIT"], lb["LBTESTCD"], lb["LBTEST"], lb["LBORRES"],
                   lb["LBORRESU"], lb["LBDTC"]])
    # VS sheet
    ws = wb.create_sheet("VS")
    ws.append(["STUDYID", "SUBJECT", "SITEID", "VISITNUM", "VSTESTCD", "VSTEST",
               "VS_RESULT", "VS_UNIT", "VISIT_DATE"])
    for vs in make_vs(subjects, center):
        ws.append([vs["STUDYID"], vs["SUBJID"], vs["SITEID"], vs["VISITNUM"],
                   vs["VSTESTCD"], vs["VSTEST"], vs["VSORRES"], vs["VSORRESU"],
                   vs["VSDTC"]])
    # AE sheet
    ws = wb.create_sheet("AE")
    ws.append(["STUDYID", "SUBJID", "SITEID", "AETERM", "AESEV", "AESER",
               "AE_START", "AE_END"])
    for ae in make_ae(subjects, center):
        ws.append([ae["STUDYID"], ae["SUBJID"], ae["SITEID"], ae["AETERM"],
                   ae["AESEV"], ae["AESER"], ae["AESTDTC"], ae["AEENDTC"]])
    wb.save(OUT / f"{center}_excel.xlsx")


# ----------------------------- CSV -----------------------------
def write_csv(center: str, subj_n: int):
    subjects = make_subjects(center, subj_n, 400)
    rows = []
    for s in subjects:
        s["DOMAIN"] = "DM"
        rows.append({**s, "VISITNUM": 1})
    for lb in make_lab(subjects, center, {"GLUC": "mg/dL"}):
        rows.append({**lb, "DOMAIN": "LB"})
    for vs in make_vs(subjects, center):
        rows.append({**vs, "DOMAIN": "VS"})
    for ae in make_ae(subjects, center):
        rows.append({**ae, "DOMAIN": "AE"})
    for ex in make_ex(subjects, center):
        rows.append({**ex, "DOMAIN": "EX"})
    fields = sorted({k for r in rows for k in r})
    with open(OUT / f"{center}_custom.csv", "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        for r in rows:
            w.writerow(r)


def write_large_csv(path: Path, n_rows: int):
    """生成大规模 CRF 数据用于性能测试。"""
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["DOMAIN", "STUDYID", "SUBJID", "SITEID", "VISITNUM",
                    "LBTESTCD", "LBORRES", "LBORRESU", "SEX", "AGE", "RFICDTC"])
        centers = ["C01", "C02", "C03", "C04", "C05", "C06", "C07", "C08"]
        units = {"GLUC": "mg/dL", "CREAT": "mg/dL", "ALT": "U/L", "HGB": "g/dL"}
        for i in range(n_rows):
            c = centers[i % len(centers)]
            code = list(units)[i % len(units)]
            val = round(random.uniform(60, 120), 1)
            w.writerow(["LB", STUDYID, f"{c}-{i:06d}", c, 1 + (i % 3),
                        code, val, units[code], "M" if i % 2 else "F",
                        random.randint(18, 80), "2024-02-01"])


if __name__ == "__main__":
    n = int(sys.argv[sys.argv.index("--large") + 1]) if "--large" in sys.argv else 0
    if n:
        write_large_csv(OUT / f"large_{n}.csv", n)
        print(f"生成大规模数据 sample/large_{n}.csv ({n} 行)")
    else:
        write_xml("C01", 12)
        write_excel("C04", 15)
        write_csv("C08", 20)
        print("生成测试数据: sample/C01_rave.xml, sample/C04_excel.xlsx, sample/C08_custom.csv")
