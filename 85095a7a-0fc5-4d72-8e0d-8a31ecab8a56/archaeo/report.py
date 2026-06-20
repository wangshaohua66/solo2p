import os
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from jinja2 import Environment, FileSystemLoader, select_autoescape

from .config import get_config_dir
from .logger import get_logger
from .models import (
    Artifact,
    ArtifactCategory,
    BudgetItem,
    Person,
    PersonRole,
    Project,
    ProjectPhase,
    ProjectStatus,
    ReportData,
    Sample,
    SampleStatus,
    SampleType,
    Stratum,
    Trench,
)
from . import db

logger = get_logger(__name__)

_TEMPLATE_DIR = Path(__file__).parent / "templates"


def _get_template_env() -> Environment:
    return Environment(
        loader=FileSystemLoader(str(_TEMPLATE_DIR)),
        autoescape=select_autoescape(["html", "xml"]),
        trim_blocks=True,
        lstrip_blocks=True,
    )


def _format_date(value: Optional[date]) -> str:
    if not value:
        return ""
    return value.strftime("%Y年%m月%d日")


def _format_datetime(value: Optional[datetime]) -> str:
    if not value:
        return ""
    return value.strftime("%Y年%m月%d日 %H:%M")


def _format_number(value: float, decimals: int = 2) -> str:
    return f"{value:,.{decimals}f}"


def _get_phase_name(phase: str) -> str:
    return ProjectPhase.get_phase_name(ProjectPhase(phase))


def _get_status_name(status: str) -> str:
    status_map = {
        "not_started": "未开始",
        "in_progress": "进行中",
        "completed": "已完成",
        "suspended": "已暂停",
    }
    return status_map.get(status, status)


def _get_category_name(cat: str) -> str:
    return ArtifactCategory.get_category_name(ArtifactCategory(cat))


def _get_sample_type_name(st: str) -> str:
    return SampleType.get_type_name(SampleType(st))


def _get_sample_status_name(st: str) -> str:
    status_map = {
        "collected": "已采集",
        "sent": "已送检",
        "testing": "检测中",
        "completed": "已完成",
        "overdue": "已超期",
    }
    return status_map.get(st, st)


def _get_role_name(role: str) -> str:
    return PersonRole.get_role_name(PersonRole(role))


def generate_briefing(project_id: int, output_path: Optional[Path] = None, 
                      format: str = "doc") -> Path:
    project = db.get_project(project_id)
    if not project:
        raise ValueError(f"项目 {project_id} 不存在")

    trenches = db.list_trenches(project_id=project_id, limit=1000)
    artifacts = db.list_artifacts(project_id=project_id, limit=10000)
    samples = db.list_samples(project_id=project_id, limit=1000)

    strata = []
    for trench in trenches:
        trench_strata = db.list_strata(trench_id=trench.id, limit=500)
        strata.extend(trench_strata)

    budget_items = db.list_budget_items(project_id=project_id)
    budget_summary = db.get_project_budget_summary(project_id)

    personnel = []
    assignments = db.list_assignments(project_id=project_id, assignment_type="person", limit=500)
    person_ids = set()
    for a in assignments:
        if a.person_id:
            person_ids.add(a.person_id)
    for pid in person_ids:
        person = db.get_person(pid)
        if person:
            personnel.append(person)

    report_data = ReportData(
        project=project,
        trenches=trenches,
        artifacts=artifacts,
        samples=samples,
        strata=strata,
        personnel=personnel,
        budget_items=budget_items,
        generated_at=datetime.now(),
    )

    env = _get_template_env()
    env.filters["format_date"] = _format_date
    env.filters["format_datetime"] = _format_datetime
    env.filters["format_number"] = _format_number
    env.filters["phase_name"] = _get_phase_name
    env.filters["status_name"] = _get_status_name
    env.filters["category_name"] = _get_category_name
    env.filters["sample_type_name"] = _get_sample_type_name
    env.filters["sample_status_name"] = _get_sample_status_name
    env.filters["role_name"] = _get_role_name

    template = env.get_template("briefing.html")

    artifact_stats = _count_by_category(artifacts)
    sample_stats = _count_by_sample_type(samples)
    sample_status_stats = _count_by_sample_status(samples)

    context = {
        "report": report_data,
        "project": project,
        "trenches": trenches,
        "artifacts": artifacts,
        "samples": samples,
        "strata": strata,
        "personnel": personnel,
        "budget_items": budget_items,
        "budget_summary": budget_summary,
        "artifact_stats": artifact_stats,
        "sample_stats": sample_stats,
        "sample_status_stats": sample_status_stats,
        "trench_count": len(trenches),
        "artifact_count": len(artifacts),
        "sample_count": len(samples),
        "stratum_count": len(strata),
        "generated_at": datetime.now(),
    }

    html_content = template.render(context)

    if format == "doc":
        html_content = _wrap_html_for_word(html_content, f"{project.name} - 发掘简报")

    if output_path is None:
        report_dir = get_config_dir() / "reports"
        report_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        ext = ".doc" if format == "doc" else ".html"
        output_path = report_dir / f"briefing_{project.code}_{timestamp}{ext}"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html_content)

    logger.info(f"发掘简报已生成: {output_path}")
    return output_path


def _wrap_html_for_word(html_content: str, title: str) -> str:
    word_header = f'''<html xmlns:o="urn:schemas-microsoft-com:office:office"
xmlns:w="urn:schemas-microsoft-com:office:word"
xmlns="http://www.w3.org/TR/REC-html40">

<head>
<meta charset="utf-8">
<meta http-equiv="Content-Type" content="application/msword; charset=utf-8">
<title>{title}</title>
<!--[if gte mso 9]><xml>
<w:WordDocument>
<w:View>Print</w:View>
<w:Zoom>100</w:Zoom>
<w:DoNotOptimizeForBrowser/>
</w:WordDocument>
</xml><![endif]-->
<style>
@page {{
    size: A4;
    margin: 2cm;
}}
body {{
    font-family: "Microsoft YaHei", "SimSun", serif;
    font-size: 12pt;
    line-height: 1.5;
}}
h1, h2, h3 {{
    font-family: "Microsoft YaHei", "SimHei", sans-serif;
}}
table {{
    border-collapse: collapse;
    width: 100%;
}}
table, th, td {{
    border: 1px solid #000;
}}
th, td {{
    padding: 6px;
    text-align: left;
}}
th {{
    background-color: #f0f0f0;
    font-weight: bold;
}}
</style>
</head>
<body>
'''

    body_start = html_content.find('<body')
    if body_start != -1:
        body_start = html_content.find('>', body_start) + 1
    else:
        body_start = 0

    body_end = html_content.find('</body>')
    if body_end == -1:
        body_end = len(html_content)

    body_content = html_content[body_start:body_end]

    word_footer = '''
</body>
</html>'''

    return word_header + body_content + word_footer


def generate_annual_report(year: int, output_path: Optional[Path] = None) -> Path:
    projects = db.list_projects(limit=1000)

    project_summaries = []
    total_budget = 0.0
    total_actual = 0.0
    total_artifacts = 0
    total_samples = 0
    total_trenches = 0

    for project in projects:
        start_year = project.start_date.year if project.start_date else None
        end_year = project.end_date.year if project.end_date else None

        if start_year and end_year:
            if start_year <= year <= end_year or (project.status == ProjectStatus.IN_PROGRESS and start_year <= year):
                pass
            elif start_year > year or (end_year and end_year < year):
                continue
        elif project.created_at and project.created_at.year == year:
            pass
        else:
            continue

        artifacts = db.list_artifacts(project_id=project.id, limit=10000)
        samples = db.list_samples(project_id=project.id, limit=1000)
        trenches = db.list_trenches(project_id=project.id, limit=1000)
        budget_summary = db.get_project_budget_summary(project.id)

        summary = {
            "project": project,
            "artifact_count": len(artifacts),
            "sample_count": len(samples),
            "trench_count": len(trenches),
            "budget_summary": budget_summary,
        }
        project_summaries.append(summary)

        total_budget += budget_summary["total_budgeted"]
        total_actual += budget_summary["total_actual"]
        total_artifacts += len(artifacts)
        total_samples += len(samples)
        total_trenches += len(trenches)

    phase_stats = _count_projects_by_phase(projects)
    status_stats = _count_projects_by_status(projects)

    persons = db.list_persons(limit=1000)
    person_role_stats = _count_persons_by_role(persons)

    env = _get_template_env()
    env.filters["format_date"] = _format_date
    env.filters["format_datetime"] = _format_datetime
    env.filters["format_number"] = _format_number
    env.filters["phase_name"] = _get_phase_name
    env.filters["status_name"] = _get_status_name
    env.filters["category_name"] = _get_category_name
    env.filters["sample_type_name"] = _get_sample_type_name
    env.filters["sample_status_name"] = _get_sample_status_name
    env.filters["role_name"] = _get_role_name

    template = env.get_template("annual_report.html")

    context = {
        "year": year,
        "project_count": len(project_summaries),
        "total_budget": total_budget,
        "total_actual": total_actual,
        "execution_rate": (total_actual / total_budget * 100) if total_budget > 0 else 0,
        "total_artifacts": total_artifacts,
        "total_samples": total_samples,
        "total_trenches": total_trenches,
        "total_personnel": len(persons),
        "projects": project_summaries,
        "phase_stats": phase_stats,
        "status_stats": status_stats,
        "person_role_stats": person_role_stats,
        "generated_at": datetime.now(),
    }

    html_content = template.render(context)

    if output_path is None:
        report_dir = get_config_dir() / "reports"
        report_dir.mkdir(parents=True, exist_ok=True)
        output_path = report_dir / f"annual_report_{year}.html"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(html_content)

    logger.info(f"年度报告已生成: {output_path}")
    return output_path


def generate_budget_report(project_id: int, quarter: Optional[int] = None,
                           year: Optional[int] = None) -> Dict[str, Any]:
    project = db.get_project(project_id)
    if not project:
        raise ValueError(f"项目 {project_id} 不存在")

    if year is None:
        year = date.today().year

    budget_items = db.list_budget_items(project_id=project_id)

    def _get_item_year(it: "BudgetItem") -> Optional[int]:
        if it.expenditure_date is not None:
            return it.expenditure_date.year
        if it.year is not None:
            return it.year
        return None

    def _get_item_quarter(it: "BudgetItem") -> Optional[int]:
        if it.quarter is not None:
            return it.quarter
        if it.expenditure_date is not None:
            return (it.expenditure_date.month - 1) // 3 + 1
        return None

    def _filter_by_year(items: List["BudgetItem"]) -> List["BudgetItem"]:
        result = []
        for it in items:
            item_year = _get_item_year(it)
            if item_year is not None and item_year != year:
                continue
            result.append(it)
        return result

    items_in_year = _filter_by_year(budget_items)

    def _build_quarterly_breakdown(items: List["BudgetItem"]) -> Dict[int, Dict[str, Any]]:
        breakdown: Dict[int, Dict[str, Any]] = {}
        for q in range(1, 5):
            breakdown[q] = {"budgeted": 0.0, "actual": 0.0, "count": 0}
        for it in items:
            q = _get_item_quarter(it)
            if q is not None and 1 <= q <= 4:
                breakdown[q]["budgeted"] += it.budgeted
                breakdown[q]["actual"] += it.actual
                breakdown[q]["count"] += 1
        for q in range(1, 5):
            breakdown[q]["budgeted"] = round(breakdown[q]["budgeted"], 2)
            breakdown[q]["actual"] = round(breakdown[q]["actual"], 2)
        return breakdown

    def _count_unassigned(items: List["BudgetItem"]) -> Dict[str, int]:
        no_year = 0
        no_quarter = 0
        for it in items:
            if _get_item_year(it) is None:
                no_year += 1
            if _get_item_quarter(it) is None:
                no_quarter += 1
        return {"no_year": no_year, "no_quarter": no_quarter}

    unassigned = _count_unassigned(items_in_year)

    if quarter and 1 <= quarter <= 4:
        filtered_items = []
        for item in items_in_year:
            item_q = _get_item_quarter(item)
            if item_q == quarter:
                filtered_items.append(item)

        total_budgeted = round(sum(it.budgeted for it in filtered_items), 2)
        total_actual = round(sum(it.actual for it in filtered_items), 2)
        execution_rate = (total_actual / total_budgeted * 100) if total_budgeted > 0 else 0.0
        deviation_items = [item for item in filtered_items if item.has_deviation]
        quarterly_breakdown = _build_quarterly_breakdown(items_in_year)

        return {
            "project_name": project.name,
            "project_code": project.code,
            "year": year,
            "quarter": quarter,
            "total_budgeted": total_budgeted,
            "total_actual": total_actual,
            "execution_rate": round(execution_rate, 2),
            "item_count": len(filtered_items),
            "deviation_count": len(deviation_items),
            "deviation_items": deviation_items,
            "budget_items": filtered_items,
            "unassigned_count": unassigned["no_quarter"],
            "year_unassigned_count": unassigned["no_year"],
            "quarterly_breakdown": quarterly_breakdown,
        }

    deviation_items = [item for item in items_in_year if item.has_deviation]
    total_budgeted = round(sum(it.budgeted for it in items_in_year), 2)
    total_actual = round(sum(it.actual for it in items_in_year), 2)
    execution_rate = (total_actual / total_budgeted * 100) if total_budgeted > 0 else 0.0
    quarterly_breakdown = _build_quarterly_breakdown(items_in_year)

    return {
        "project_name": project.name,
        "project_code": project.code,
        "year": year,
        "quarter": quarter,
        "total_budgeted": total_budgeted,
        "total_actual": total_actual,
        "execution_rate": round(execution_rate, 2),
        "item_count": len(items_in_year),
        "deviation_count": len(deviation_items),
        "deviation_items": deviation_items,
        "budget_items": items_in_year,
        "unassigned_count": unassigned["no_quarter"],
        "year_unassigned_count": unassigned["no_year"],
        "quarterly_breakdown": quarterly_breakdown,
    }


def export_budget_excel(project_id: int, output_path: Path, 
                        quarter: Optional[int] = None, 
                        year: Optional[int] = None) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        logger.warning("openpyxl 未安装，无法导出 Excel")
        raise

    budget_data = generate_budget_report(project_id, quarter=quarter, year=year)
    budget_items = budget_data["budget_items"]

    wb = Workbook()
    ws = wb.active
    title_suffix = f" - {budget_data['year']}年第{budget_data['quarter']}季度" if budget_data.get("quarter") else f" - {budget_data['year']}年"
    ws.title = f"经费执行情况"

    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    deviation_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    q4_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    period_text = f"{budget_data['year']}年"
    if budget_data.get("quarter"):
        period_text += f"第{budget_data['quarter']}季度"

    ws["A1"] = f"{budget_data['project_name']} - 经费执行表 ({period_text})"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:I1")
    ws["A1"].alignment = center_align

    headers = ["ID", "预算科目", "年份", "季度", "支出日期", "预算金额", "实际支出", "执行率", "偏差预警"]
    col_count = len(headers)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for row_idx, item in enumerate(budget_items, 4):
        year_val = item.year if item.year else (
            item.expenditure_date.year if item.expenditure_date else "-"
        )
        q_str = f"Q{item.quarter}" if item.quarter else (
            f"Q{((item.expenditure_date.month - 1) // 3 + 1)}" if item.expenditure_date else "-"
        )
        date_str = item.expenditure_date.isoformat() if item.expenditure_date else "-"
        ws.cell(row=row_idx, column=1, value=item.id)
        ws.cell(row=row_idx, column=2, value=item.category)
        ws.cell(row=row_idx, column=3, value=year_val)
        ws.cell(row=row_idx, column=4, value=q_str)
        ws.cell(row=row_idx, column=5, value=date_str)
        ws.cell(row=row_idx, column=6, value=item.budgeted)
        ws.cell(row=row_idx, column=7, value=item.actual)
        rate_cell = ws.cell(row=row_idx, column=8, value=f"{item.execution_rate:.1f}%")

        deviation_cell = ws.cell(row=row_idx, column=9)
        row_fill = None
        if item.has_deviation:
            deviation_cell.value = "⚠️ 偏差超20%"
            deviation_cell.font = Font(color="FF0000", bold=True)
            row_fill = deviation_fill
        else:
            deviation_cell.value = "正常"

        for col in range(1, col_count + 1):
            c = ws.cell(row=row_idx, column=col)
            c.border = thin_border
            if row_fill:
                c.fill = row_fill
            c.alignment = Alignment(vertical="center")
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row_idx, column=3).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row_idx, column=4).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row_idx, column=5).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row_idx, column=6).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=row_idx, column=7).alignment = Alignment(horizontal="right", vertical="center")
        rate_cell.alignment = Alignment(horizontal="center", vertical="center")

    summary_row = len(budget_items) + 5
    ws.cell(row=summary_row, column=1, value="合计").font = Font(bold=True)
    ws.merge_cells(f"A{summary_row}:D{summary_row}")
    ws.cell(row=summary_row, column=6, value=budget_data["total_budgeted"]).font = Font(bold=True)
    ws.cell(row=summary_row, column=7, value=budget_data["total_actual"]).font = Font(bold=True)
    rate_cell = ws.cell(row=summary_row, column=8, value=f"{budget_data['execution_rate']:.1f}%")
    rate_cell.font = Font(bold=True)
    rate_cell.alignment = Alignment(horizontal="center")

    breakdown_start = summary_row + 3
    ws.cell(row=breakdown_start, column=1, value=f"{budget_data['year']}年各季度执行情况汇总").font = Font(bold=True, size=12)
    ws.merge_cells(f"A{breakdown_start}:I{breakdown_start}")
    ws.cell(row=breakdown_start, column=1).alignment = center_align

    bh_row = breakdown_start + 1
    bh_headers = ["季度", "预算金额", "实际支出", "执行率", "记录数", "占全年预算", "占全年支出"]
    for col, h in enumerate(bh_headers, 1):
        c = ws.cell(row=bh_row, column=col, value=h)
        c.font = header_font_white
        c.fill = header_fill
        c.alignment = center_align
        c.border = thin_border

    breakdown = budget_data.get("quarterly_breakdown", {})
    year_total_budgeted = budget_data["total_budgeted"]
    year_total_actual = budget_data["total_actual"]
    highlight_q = budget_data.get("quarter")

    for q in range(1, 5):
        r = bh_row + q
        qdata = breakdown.get(q, {"budgeted": 0.0, "actual": 0.0, "count": 0})
        q_rate = (qdata["actual"] / qdata["budgeted"] * 100) if qdata["budgeted"] else 0.0
        pct_budget = (qdata["budgeted"] / year_total_budgeted * 100) if year_total_budgeted else 0.0
        pct_actual = (qdata["actual"] / year_total_actual * 100) if year_total_actual else 0.0

        fill = q4_fill if highlight_q == q else None
        values = [
            f"Q{q}",
            qdata["budgeted"],
            qdata["actual"],
            f"{q_rate:.1f}%",
            qdata["count"],
            f"{pct_budget:.1f}%",
            f"{pct_actual:.1f}%",
        ]
        for col, v in enumerate(values, 1):
            c = ws.cell(row=r, column=col, value=v)
            c.border = thin_border
            c.alignment = Alignment(vertical="center")
            if col in (1, 4, 5, 6, 7):
                c.alignment = Alignment(horizontal="center", vertical="center")
            if col in (2, 3):
                c.alignment = Alignment(horizontal="right", vertical="center")
            if fill:
                c.fill = fill

    note_row = bh_row + 6
    note_idx = 0
    if budget_data.get("unassigned_count", 0) > 0:
        ws.cell(row=note_row, column=1,
                value=f"⚠️ 有 {budget_data['unassigned_count']} 条记录未指定季度/日期，不计入季度聚合")
        ws.cell(row=note_row, column=1).font = Font(color="C00000", italic=True)
        ws.merge_cells(f"A{note_row}:I{note_row}")
        note_idx += 1

    if budget_data.get("year_unassigned_count", 0) > 0:
        ws.cell(row=note_row + note_idx, column=1,
                value=f"ℹ️ 有 {budget_data['year_unassigned_count']} 条记录未指定年份，默认计入目标年份")
        ws.cell(row=note_row + note_idx, column=1).font = Font(color="7030A0", italic=True)
        ws.merge_cells(f"A{note_row + note_idx}:I{note_row + note_idx}")
        note_idx += 1

    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 13
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 15
    ws.column_dimensions["H"].width = 11
    ws.column_dimensions["I"].width = 16

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info(f"经费报表已导出: {output_path}")
    return output_path


def _count_by_category(artifacts: List[Artifact]) -> Dict[str, int]:
    stats: Dict[str, int] = {}
    for artifact in artifacts:
        cat = artifact.category.value
        stats[cat] = stats.get(cat, 0) + artifact.quantity
    return stats


def _count_by_sample_type(samples: List[Sample]) -> Dict[str, int]:
    stats: Dict[str, int] = {}
    for sample in samples:
        st = sample.sample_type.value
        stats[st] = stats.get(st, 0) + 1
    return stats


def _count_by_sample_status(samples: List[Sample]) -> Dict[str, int]:
    stats: Dict[str, int] = {}
    for sample in samples:
        st = sample.status.value
        stats[st] = stats.get(st, 0) + 1
    return stats


def _count_projects_by_phase(projects: List[Project]) -> Dict[str, int]:
    stats: Dict[str, int] = {}
    for project in projects:
        phase = project.phase.value
        stats[phase] = stats.get(phase, 0) + 1
    return stats


def _count_projects_by_status(projects: List[Project]) -> Dict[str, int]:
    stats: Dict[str, int] = {}
    for project in projects:
        status = project.status.value
        stats[status] = stats.get(status, 0) + 1
    return stats


def _count_persons_by_role(persons: List[Person]) -> Dict[str, int]:
    stats: Dict[str, int] = {}
    for person in persons:
        role = person.role.value
        stats[role] = stats.get(role, 0) + 1
    return stats
