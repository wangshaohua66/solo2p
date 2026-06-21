from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    _HAS_MPL = True
except Exception:
    _HAS_MPL = False

from logger import get_logger
from database import DatabaseManager
from config import BASE_DIR


logger = get_logger("report")
REPORT_DIR = BASE_DIR / "data" / "reports"
REPORT_DIR.mkdir(parents=True, exist_ok=True)


_HEADER_FONT = Font(name="微软雅黑", bold=True, size=12, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="4472C4")
_CELL_FONT = Font(name="微软雅黑", size=11)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_THIN = Side(style="thin", color="B4B4B4")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _style_header(ws, row: int, max_col: int) -> None:
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _BORDER


def _style_data(ws, start_row: int, end_row: int, max_col: int) -> None:
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = _CELL_FONT
            cell.alignment = _CENTER
            cell.border = _BORDER


def _auto_width(ws, data: List[List]) -> None:
    for col_idx in range(1, len(data[0]) + 1 if data else 2):
        max_len = 10
        for row in data:
            if len(row) >= col_idx:
                val = str(row[col_idx - 1]) if row[col_idx - 1] is not None else ""
                max_len = max(max_len, min(len(val) * 2, 40))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len


def generate_fair_report(fair_id: str, output_path: Optional[str] = None) -> str:
    db = DatabaseManager()
    fair_rows = db.query_all("SELECT * FROM job_fairs WHERE fair_id=?", (fair_id,))
    if not fair_rows:
        raise ValueError(f"招聘会不存在: {fair_id}")
    fair = dict(fair_rows[0])

    if not output_path:
        safe_title = "".join(c for c in fair["title"] if c.isalnum() or c in "_-")[:30]
        output_path = str(REPORT_DIR / f"fair_{fair_id}_{safe_title}.xlsx")

    wb = Workbook()

    ws_info = wb.active
    ws_info.title = "招聘会概况"
    info_data = [
        ["项目", "内容"],
        ["招聘会名称", fair.get("title", "")],
        ["举办日期", fair.get("fair_date", "")],
        ["举办地点", fair.get("location", "")],
        ["主办单位", fair.get("organizer", "")],
        ["企业数量", fair.get("company_count", 0)],
        ["数据来源", fair.get("site_name", "")],
        ["采集时间", fair.get("crawl_time", "")],
    ]
    for r, row in enumerate(info_data, 1):
        for c, val in enumerate(row, 1):
            ws_info.cell(row=r, column=c, value=val)
    _style_header(ws_info, 1, 2)
    _style_data(ws_info, 2, len(info_data), 2)
    _auto_width(ws_info, info_data)

    companies = [dict(r) for r in db.query_companies_by_fair(fair_id)]
    ws_comp = wb.create_sheet("参展企业")
    comp_headers = ["序号", "企业名称", "展位号", "行业", "企业类型", "联系人", "联系电话", "邮箱"]
    comp_data = [comp_headers]
    for i, c in enumerate(companies, 1):
        comp_data.append([
            i, c.get("name", ""), c.get("booth_number", ""), c.get("industry", ""),
            c.get("company_type", ""), c.get("contact_person", ""),
            c.get("contact_phone", ""), c.get("contact_email", ""),
        ])
    for r, row in enumerate(comp_data, 1):
        for col_idx, val in enumerate(row, 1):
            ws_comp.cell(row=r, column=col_idx, value=val)
    _style_header(ws_comp, 1, len(comp_headers))
    _style_data(ws_comp, 2, len(comp_data), len(comp_headers))
    _auto_width(ws_comp, comp_data)

    jobs = [dict(r) for r in db.query_jobs_by_fair(fair_id)]
    ws_jobs = wb.create_sheet("招聘岗位")
    job_headers = ["序号", "企业名称", "岗位名称", "学历要求", "专业要求", "薪资下限", "薪资上限", "薪资单位", "工作地点", "招聘人数"]
    job_data = [job_headers]
    comp_map = {c["company_id"]: c.get("name", "") for c in companies}
    for i, j in enumerate(jobs, 1):
        job_data.append([
            i, comp_map.get(j.get("company_id", ""), ""), j.get("title", ""),
            j.get("education", ""), j.get("major", ""),
            j.get("salary_min") or "", j.get("salary_max") or "",
            j.get("salary_unit", ""), j.get("location", ""), j.get("job_count", 1),
        ])
    for r, row in enumerate(job_data, 1):
        for col_idx, val in enumerate(row, 1):
            ws_jobs.cell(row=r, column=col_idx, value=val)
    _style_header(ws_jobs, 1, len(job_headers))
    _style_data(ws_jobs, 2, len(job_data), len(job_headers))
    _auto_width(ws_jobs, job_data)

    ws_stats = wb.create_sheet("统计分析")
    edu_stats: Dict[str, int] = {}
    major_stats: Dict[str, int] = {}
    salary_ranges = {"5k以下": 0, "5k-8k": 0, "8k-12k": 0, "12k-20k": 0, "20k以上": 0}
    for j in jobs:
        edu = j.get("education") or "不限"
        edu_stats[edu] = edu_stats.get(edu, 0) + 1
        major = j.get("major") or "不限"
        if major and len(major) < 20:
            major_stats[major] = major_stats.get(major, 0) + 1
        avg_sal = ((j.get("salary_min") or 0) + (j.get("salary_max") or 0)) / 2
        if avg_sal == 0:
            pass
        elif avg_sal < 5000:
            salary_ranges["5k以下"] += 1
        elif avg_sal < 8000:
            salary_ranges["5k-8k"] += 1
        elif avg_sal < 12000:
            salary_ranges["8k-12k"] += 1
        elif avg_sal < 20000:
            salary_ranges["12k-20k"] += 1
        else:
            salary_ranges["20k以上"] += 1

    stats_data = [["维度", "类别", "数量"]]
    stats_data.append(["学历要求", "", ""])
    for k, v in sorted(edu_stats.items(), key=lambda x: -x[1]):
        stats_data.append(["", k, v])
    stats_data.append(["薪资分布", "", ""])
    for k, v in salary_ranges.items():
        stats_data.append(["", k, v])
    stats_data.append(["热门专业(Top10)", "", ""])
    for k, v in sorted(major_stats.items(), key=lambda x: -x[1])[:10]:
        stats_data.append(["", k, v])

    for r, row in enumerate(stats_data, 1):
        for col_idx, val in enumerate(row, 1):
            ws_stats.cell(row=r, column=col_idx, value=val)
    _style_header(ws_stats, 1, 3)
    _style_data(ws_stats, 2, len(stats_data), 3)
    _auto_width(ws_stats, stats_data)

    wb.save(output_path)
    logger.info(f"招聘会报表已生成: {output_path}")
    return output_path


def generate_charts(fair_id: str, output_dir: Optional[str] = None) -> List[str]:
    if not _HAS_MPL:
        logger.warning("matplotlib未安装，跳过图表生成")
        return []

    db = DatabaseManager()
    jobs = [dict(r) for r in db.query_jobs_by_fair(fair_id)]
    if not jobs:
        return []

    output_dir = Path(output_dir) if output_dir else REPORT_DIR
    output_dir.mkdir(parents=True, exist_ok=True)
    paths = []

    plt.rcParams["font.sans-serif"] = ["Arial Unicode MS", "SimHei", "DejaVu Sans"]
    plt.rcParams["axes.unicode_minus"] = False

    edu_stats: Dict[str, int] = {}
    for j in jobs:
        edu = j.get("education") or "不限"
        edu_stats[edu] = edu_stats.get(edu, 0) + 1
    if edu_stats:
        fig, ax = plt.subplots(figsize=(8, 5))
        items = sorted(edu_stats.items(), key=lambda x: -x[1])
        ax.bar([k for k, _ in items], [v for _, v in items], color="#4472C4")
        ax.set_title("岗位学历要求分布")
        ax.set_ylabel("岗位数量")
        plt.tight_layout()
        p = str(output_dir / f"fair_{fair_id}_education.png")
        fig.savefig(p, dpi=150)
        plt.close(fig)
        paths.append(p)

    salary_ranges = {"5k以下": 0, "5k-8k": 0, "8k-12k": 0, "12k-20k": 0, "20k以上": 0}
    for j in jobs:
        avg_sal = ((j.get("salary_min") or 0) + (j.get("salary_max") or 0)) / 2
        if avg_sal == 0:
            continue
        elif avg_sal < 5000:
            salary_ranges["5k以下"] += 1
        elif avg_sal < 8000:
            salary_ranges["5k-8k"] += 1
        elif avg_sal < 12000:
            salary_ranges["8k-12k"] += 1
        elif avg_sal < 20000:
            salary_ranges["12k-20k"] += 1
        else:
            salary_ranges["20k以上"] += 1
    if any(v > 0 for v in salary_ranges.values()):
        fig, ax = plt.subplots(figsize=(8, 5))
        labels = list(salary_ranges.keys())
        values = list(salary_ranges.values())
        colors = ["#70AD47", "#FFC000", "#ED7D31", "#4472C4", "#C00000"]
        ax.pie(values, labels=labels, autopct="%1.1f%%", colors=colors, startangle=90)
        ax.set_title("岗位薪资分布")
        plt.tight_layout()
        p = str(output_dir / f"fair_{fair_id}_salary.png")
        fig.savefig(p, dpi=150)
        plt.close(fig)
        paths.append(p)

    logger.info(f"图表生成完成: {len(paths)}张")
    return paths


def generate_summary_report(output_path: Optional[str] = None) -> str:
    db = DatabaseManager()
    if not output_path:
        output_path = str(REPORT_DIR / f"summary_{datetime.now().strftime('%Y%m%d')}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "数据总览"

    total_fairs = db.query_all("SELECT COUNT(*) AS c FROM job_fairs")[0]["c"]
    total_companies = db.query_all("SELECT COUNT(*) AS c FROM companies")[0]["c"]
    total_jobs = db.query_all("SELECT COUNT(*) AS c FROM jobs")[0]["c"]
    total_submissions = db.query_all("SELECT COUNT(*) AS c FROM submissions")[0]["c"]

    overview = [
        ["指标", "数量"],
        ["招聘会总数", total_fairs],
        ["企业总数", total_companies],
        ["岗位总数", total_jobs],
        ["简历投递总数", total_submissions],
    ]
    for r, row in enumerate(overview, 1):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    _style_header(ws, 1, 2)
    _style_data(ws, 2, len(overview), 2)
    _auto_width(ws, overview)

    status_rows = db.query_all(
        "SELECT status, COUNT(*) AS c FROM submissions GROUP BY status ORDER BY c DESC"
    )
    ws2 = wb.create_sheet("投递状态分布")
    status_data = [["投递状态", "数量"]]
    from status_tracker import STATUS_LABELS
    for row in status_rows:
        status_data.append([STATUS_LABELS.get(row["status"], row["status"]), row["c"]])
    for r, row in enumerate(status_data, 1):
        for c, val in enumerate(row, 1):
            ws2.cell(row=r, column=c, value=val)
    _style_header(ws2, 1, 2)
    _style_data(ws2, 2, len(status_data), 2)
    _auto_width(ws2, status_data)

    wb.save(output_path)
    logger.info(f"汇总报表已生成: {output_path}")
    return output_path
